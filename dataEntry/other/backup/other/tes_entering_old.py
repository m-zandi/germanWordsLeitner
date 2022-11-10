import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 
# import pandas as pd
import openpyxl
import pymongo
import string
from bson.objectid import ObjectId
import bson
from tkinter import *
from tkinter import ttk


# from xlrd import open_workbook as OW

class Db:
    def __init__(self):
        self.connection = pymongo.MongoClient("localhost",27017)
        self.database= self.connection["woerterbuch"]
        self.userCollection = self.database["user"]
        self.wordCollection = self.database["word"]
        self.meaningCollection = self.database["meaning"]
        self.bookCollection = self.database["book"]
        self.sectionCollection =  self.database["section"]
        self.settingCollection = self.database["setting"]
        self.backupCollection = self.database["backup"]
        self.testMeaningCollection  = self.database["testMeaning"]

class Finding:
    def __init__(self):
        # wordIdTitleEx bookIdTitleEx meaningIdTitleEx
        self.obj_Db = Db()
        self.wordNumTitleEx = "word_number"
        ## Word Filed
        self.wordIdTitleEx = "word id"
        self.wordTitleEx = "word"
        self.wordTypeTitleEx = "word Type"
        self.articleTitleEx = "article"
        self.nWFileAudioTitleEx = "name_wort_file_Audio"
        self.aTWTitleEx = "address_Telegram_wort"
        wordFiled = [self.wordIdTitleEx,self.wordTitleEx,self.wordTypeTitleEx,self.articleTitleEx,self.nWFileAudioTitleEx,self.aTWTitleEx]
        ## Book Field
        self.bookIdTitleEx = "book id"
        self.bookNTitleEx = "book name"
        self.chapTitleEx = "chapter"
        self.contTitleEx = "content"
        self.pageTitleEx = "page"
        self.pubTitleEx = "publisher"
        bookField = [self.bookIdTitleEx,self.bookNTitleEx,self.chapTitleEx,self.contTitleEx,self.pageTitleEx,self.pubTitleEx]
        ## Meaning field
        self.meaningIdTitleEx = "meaning id"
        self.mNumDafTitleEx = "meaning number(DAF)"
        self.deuTitleEx = "deutsch"
        self.engTitleEx = "english"
        self.synTitleEx ="Synonym"
        self.perTitleEx = "persian"
        self.nDeFileAuTitleEx = "name_De_file Audio"
        self.addTelDeuTitleEx = "address_Telegram_Deutsch"

        self.naSynFileAuEx = "name_Synonym_file Audio"
        self.addTelSynTitleEx = "address_Telegram_Synonym"

        self.naEngFileAuTitleEx = "name_English_file Audio"
        self.addTelEngTitleEx="address_Telegram_English"

        self.naPerFileAudioTitleEx="name_Persian_file Audio"
        self.addTelPer = "address_Telegram_Persian"

        optionalMeaningFiled = [self.nDeFileAuTitleEx,self.addTelDeuTitleEx,self.naSynFileAuEx,self.addTelSynTitleEx,self.naEngFileAuTitleEx,self.addTelEngTitleEx]
        meaningField=[self.meaningIdTitleEx,self.mNumDafTitleEx,self.deuTitleEx,self.engTitleEx,self.synTitleEx,self.perTitleEx] + optionalMeaningFiled
        ##
        self.exTitles = [self.wordNumTitleEx] + wordFiled + bookField + meaningField
        self.path = r"D:\project\WoerterbuchProject\mainV2\dataEntry\words.xlsx" 
        self.sheetName = "orginal"

    def findWordNDeutschInExcel(self,word,deutsch,path,sheetName):
        wordTitle = self.wordTitleEx
        deutschTitle = self.deuTitleEx
        mySheet = self.getSheet(path,sheetName)
        wordColumn,deutschColumn,deutschRow,wordRow,output = None,None,[],[],[]
        verificationDeutsch = False
        for row in range(1,mySheet.max_row+1):
            for col in range(1,mySheet.max_column+1):
                cell_value = mySheet.cell(row,col).value
                if wordTitle == cell_value:
                    wordColumn = col
                elif deutschTitle == cell_value:
                    deutschColumn = col
                elif deutsch == cell_value:
                    deutschRow.append(row)
                    verificationDeutsch=True
                elif word == cell_value:
                    wordRow.append(row)
        intersection = set(deutschRow)&set(wordRow)
        for eachRow in intersection:
            temp = []
            temp.append(eachRow)
            temp.append(wordColumn)
            output.append(tuple(temp))
        if verificationDeutsch is False:
            output = [False]
        return output


    def getSheetNWorkbook(self,path,sheet):
        wb = load_workbook(filename=path)
        mySheet = wb.get_sheet_by_name(sheet)
        return mySheet,wb
    def getSheet(self,path,sheet):
        wb = load_workbook(filename=path)
        mySheet = wb.get_sheet_by_name(sheet)
        return mySheet

    def getWordsMongo(self):
        output = []
        for col in self.obj_Db.wordCollection.find({},{"word":1,"_id":0}):
            for keys in col.keys():
                output.append(col[keys])
        return output
    def findValueInExcel(self,val,pathFile,sheetName =None):
        output = []
        wb = openpyxl.load_workbook(pathFile)
        if sheetName is None:
            sheetsName = wb.sheetnames
            for eachSheet in sheetsName:
                mySheet = wb.get_sheet_by_name(eachSheet)
                    
                for row in range(1,mySheet.max_row+1):
                    # print(f"{row = }")
                    for col in range(1,mySheet.max_column+1):
                        # print(f"{col = }")
                        tempList = []
                        cell_value = mySheet.cell(row,col).value
                        # cell_value = mySheet.cell(row = row,col=col).value
                        if cell_value == val:
                            tempList.append(eachSheet)
                            tempList.append(col)
                            tempList.append(row)
                            tupleCon = tuple(tempList)
                            output.append(tupleCon)
        else:
            mySheet = wb.get_sheet_by_name(sheetName)        
            for row in range(1,mySheet.max_row+1):
                # print(f"{row = }")
                for col in range(1,mySheet.max_column+1):
                    # print(f"{col = }")
                    tempList = []
                    cell_value = mySheet.cell(row,col).value
                    # cell_value = mySheet.cell(row = row,col=col).value
                    if cell_value == val:
                        tempList.append(sheetName)
                        tempList.append(col)
                        tempList.append(row)
                        tupleCon = tuple(tempList)
                        output.append(tupleCon)

        # print(f"findValueInExcel {output = }")
        #sheetName,col,row
        if  len(output) ==0 :
            
            output=False
        return output


    
    def getExcelWords(self,pathNName,sheetName):
        idNwords = []
        words = []
        wordNWordIdNBookIdNmeaningId = []
        sheet = self.getSheet(pathNName,sheetName)
        wordNumSheetColRow = self.findValueInExcel(self.wordNumTitleEx,pathNName,sheetName=sheetName)
        # print(f"{wordNumSheetColRow = }")
        columnWordNum =wordNumSheetColRow[0][1] 
        columnAlphaWordNum = openpyxl.utils.cell.get_column_letter(columnWordNum)

        wordTitleSheetColRow = self.findValueInExcel(self.wordTitleEx,pathNName,sheetName=sheetName)
        # print(f"{wordTitleSheetColRow = }")
        columnWordTitle =wordTitleSheetColRow[0][1] 
        columnAlphaWordTitle = openpyxl.utils.cell.get_column_letter(columnWordTitle)
        rowStart = wordTitleSheetColRow[0][2]+1



        # print(f"{columnAlphaWordTitle = }")
        for i in range(rowStart,len(sheet[columnAlphaWordTitle])+1):
            tempList = []
            tempList.append(sheet[f"{columnAlphaWordNum}{i}"].value)
            tempList.append(sheet[f"{columnAlphaWordTitle}{i}"].value)

            idNwords.append(tuple(tempList))
            words.append(sheet[f"{columnAlphaWordTitle}{i}"].value)
        return idNwords,words
    
    def getWordsNWordIdsNBookIdsNmeaningIds(self,pathNName,sheetName):
        output = []
        _,words = self.getExcelWords(pathNName,sheetName)
        mySheet = self.getSheet(pathNName,sheetName)
        wordIdSheetColRow = self.findValueInExcel(self.wordIdTitleEx ,pathNName,sheetName=sheetName)
        wordIdColumn = wordIdSheetColRow[0][1]
        bookIdSheetColRow = self.findValueInExcel(self.bookIdTitleEx ,pathNName,sheetName=sheetName)
        bookIdColumn =  bookIdSheetColRow[0][1]
        meaningIdSheetColRow = self.findValueInExcel(self.meaningIdTitleEx ,pathNName,sheetName=sheetName)
        meaningIdColumn =  meaningIdSheetColRow[0][1]
 
        for eachWord in words:
  
            listInside = []
            listInside.append(eachWord)
            wordSheetColRow = self.findValueInExcel(eachWord ,pathNName,sheetName=sheetName)
            for eachOne in wordSheetColRow:
                row = eachOne[2]
                # print(f"{row = }\n{wordIdColumn = }")
                wordId = mySheet.cell(row,wordIdColumn).value
                listInside.append(wordId)

                # print(f"{row = }\n{meaningIdColumn = }")
                meaningId = mySheet.cell(row,meaningIdColumn).value
                listInside.append(meaningId)

                # print(f"{row = }\n{bookIdColumn = }")
                bookId = mySheet.cell(row,bookIdColumn).value
                listInside.append(bookId)
                output.append(tuple(listInside))
        return output








    def getWordsSeparated(self,idNwords):
        numbers = []
        words = []
        for i in range(len(idNwords)):
            numbers.append(idNwords[i][0])
            words.append(idNwords[i][1])
        return numbers,words

    def getMongoWords(self):
        mongoWords = []
        for col in self.obj_Db.wordCollection.find({},{"word":1,"_id":0}):
            for keys in col.keys():
                mongoWords.append(col[keys])
        return mongoWords  

    def getmeaningIdFromExcel(self,word):
        output = []
        listOrFalse = self.findValueInExcel(word,self.path,self.sheetName)
        meaningIdTitleTup = self.findValueInExcel(self.meaningIdTitleEx,self.path,self.sheetName)
        column = meaningIdTitleTup[0][1]
        mySheet,_=self.getSheetNWorkbook(self.path,self.sheetName)
        
        if listOrFalse is not False:
            for x in listOrFalse:
                row = x[2]
                output.append(mySheet.cell(row,column).value)
        return output

    def getmeaningIdFromMongo(self,word):
        output = []
        for col in self.obj_Db.meaningCollection.find({"word":word},{"_id":1}):
            for keys in col.keys():
                output.append(col[keys])
        return output

    def getMeaningWordFromMongoAllDetails(self,word,id):
        id = ObjectId(id)
        condition = {"word":word,"_id":id}
        outputMongo = {"_id":1,"deutsch":1,"english":1,"synonym":1,"persian":1,"meaningNumDaf":1,"word":1,"synVoice":1,"engVoice":1,"perVoice":1,"deuVoice":1}
        meaningId,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer = None,None,None,None,None,None,None,None,None,None,None,None,None,None
        for col in self.obj_Db.meaningCollection.find(condition,outputMongo):
            for field,val in col.items():
                if field == "deutsch":
                    deutsch=val
                if field == "english":
                    english = val
                if field == "synonym":
                    synonym = val
                if field == "persian":
                    persian = val
                if field == "meaningNumDaf":
                    meaningNumDAF = val
                if field == "_id":
                    meaningId = val
                if field ==  "deuVoice":
                        nameDeAudio = val["deuName"]
                        addTelDeu = val["deuTelegramLink"]
                if field == "engVoice":
                        nameEnAudio = val["engName"]
                        addTelEn = val["engTelegramLink"]
                if field == "synVoice":
                        nameSynAudio = val["synName"]
                        addTelSyn = val["synTelegramLink"]     
                if field == "perVoice":
                        namePerAudio = val["perName"]
                        addTelPer = val["perTelegramLink"]

        return meaningId,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer


    def nonExistanceWordsInMongo(self,pathNName,sheetName):
        idNwords,_= self.getExcelWords(pathNName,sheetName)
        _,excelWords = self.getWordsSeparated(idNwords)
        output = False
        nonWordInMongo = []
        number = []
        mongoWords = self.getMongoWords()
        for i in excelWords:
            if i not in mongoWords:
                nonWordInMongo.append(i)
        if len(nonWordInMongo) !=0:
            output=nonWordInMongo.copy()
        return output
 
    def getFilledWordsFromExcelAllContent(self,wordVal):
        # path = r"D:\project\WoerterbuchProject\mainV2\dataEntry\words.xlsx" 
        # sheetName = "orginal"

        mySheet = self.getSheet(self.path,self.sheetName)
        path = self.path
        sheetName = self.sheetName
        output = ""
        # #sheetName,col,row
        wordValSheetCoRow = self.findValueInExcel(wordVal,path,sheetName =sheetName)
        if wordValSheetCoRow is not False:

            wordSheetCoRow =self.findValueInExcel(self.wordTitleEx,path,sheetName =sheetName)
            wordColAlpha = openpyxl.utils.cell.get_column_letter(wordSheetCoRow[0][1])

            wordTypeSheetCoRow =self.findValueInExcel(self.wordTypeTitleEx,path,sheetName =sheetName)
            articleSheetCoRow =self.findValueInExcel(self.articleTitleEx,path,sheetName =sheetName)
            voiceNameSheetCoRow =self.findValueInExcel(self.nWFileAudioTitleEx,path,sheetName =sheetName)
            voiceTelLinkSheetCoRow =self.findValueInExcel(self.aTWTitleEx,path,sheetName =sheetName)
            row = wordValSheetCoRow[0][2]

            wordType,article,voiceName,voiceTelLink = "","","",""
            wordValue = mySheet.cell(row,wordSheetCoRow[0][1]).value
            wordType =  mySheet.cell(row,wordTypeSheetCoRow[0][1]).value
            article = mySheet.cell(row,articleSheetCoRow[0][1]).value
            voiceName = mySheet.cell(row,voiceNameSheetCoRow[0][1]).value
            voiceTelLink = mySheet.cell(row,voiceTelLinkSheetCoRow[0][1]).value
            if wordValue is not None and wordType is not None and article is not None and voiceName is not None and voiceTelLink is not None:

                return wordType,article,voiceName,voiceTelLink
            else:
                return False
        else:
            return False

 
    def getWordsFromExcelAllContent(self,wordVal):
        # path = r"D:\project\WoerterbuchProject\mainV2\dataEntry\words.xlsx" 
        # sheetName = "orginal"
        mySheet = self.getSheet(self.path,self.sheetName)
        path = self.path
        sheetName = self.sheetName
        output = ""
        # #sheetName,col,row
        wordValSheetCoRow = self.findValueInExcel(wordVal,path,sheetName =sheetName)
        if wordValSheetCoRow is not False:

        
            wordSheetCoRow =self.findValueInExcel(self.wordTitleEx,path,sheetName =sheetName)
            # wordColAlpha = openpyxl.utils.cell.get_column_letter(wordSheetCoRow[0][1])

            wordTypeSheetCoRow =self.findValueInExcel(self.wordTypeTitleEx,path,sheetName =sheetName)
            articleSheetCoRow =self.findValueInExcel(self.articleTitleEx,path,sheetName =sheetName)
            voiceNameSheetCoRow =self.findValueInExcel(self.nWFileAudioTitleEx,path,sheetName =sheetName)
            voiceTelLinkSheetCoRow = self.findValueInExcel(self.aTWTitleEx,path,sheetName =sheetName)

            meaningIdSheetCoRow = self.findValueInExcel(self.meaningIdTitleEx,path,sheetName =sheetName)
            meainigidColumn = meaningIdSheetCoRow[0][1]
            row = wordValSheetCoRow[0][2]



            wordType,article,voiceName,voiceTelLink,meaningId = "","","","",[]
            wordValue = mySheet.cell(row,wordSheetCoRow[0][1]).value
            wordType =  mySheet.cell(row,wordTypeSheetCoRow[0][1]).value
            article = mySheet.cell(row,articleSheetCoRow[0][1]).value
            voiceName = mySheet.cell(row,voiceNameSheetCoRow[0][1]).value
            voiceTelLink = mySheet.cell(row,voiceTelLinkSheetCoRow[0][1]).value
            for x in range(len(wordValSheetCoRow)):
                meaningId.append(mySheet.cell(wordValSheetCoRow[x][2],meainigidColumn).value)
            return wordType,article,voiceName,voiceTelLink,meaningId
        else:
            return False
    def getWordsFromMongoAllContent(self,word):
        output = []
        outputMongo = {"wordType":1,"article":1,"voice":1,"meaningId":1,"_id":0}
        wordType,article,voiceName,voiceTelLink,meaningId=None,None,None,None,None
        for col in self.obj_Db.wordCollection.find({"word":word},outputMongo):
            for field,val in col.items():
                if field == "wordType":
                    wordType = val
                    # output.append(wordType)
                elif field == "article":
                    article = val
                    # output.append(article)
                elif field == "voice":
                    voiceName = val["name"]
                    # output.append(voiceName)
                    voiceTelLink = val["telegramLink"]
                    # output.append(voiceTelLink)
                elif field == "meaningId":
                    meaningId = val
                    # output.append(meaningId)
        output = [wordType,article,voiceName,voiceTelLink,meaningId]
    #    wordType,article,voiceName,voiceTelLink,meaningId
        return output

    def getMeaningWordFromExcelAllDetails(self,word,deutschMeaning=None,meaningId=None):
        output = []
        mySheet = self.getSheet(self.path,self.sheetName)
        path = self.path
        sheetName = self.sheetName
        row = None
        if self.findValueInExcel(word,path,sheetName =sheetName):
            # row = ""
            # print(f"{row = }")
            if deutschMeaning ==None and meaningId==None or (self.findValueInExcel(deutschMeaning,path,sheetName =sheetName) is False) or (self.findValueInExcel(meaningId,path,sheetName =sheetName) is False):
                wordValSheetCoRow = self.findValueInExcel(word,path,sheetName =sheetName)
                # print(f"{wordValSheetCoRow = }")
                # print(f"{row = }")
                row = wordValSheetCoRow[0][2]


            elif deutschMeaning !=None and meaningId ==None and (self.findValueInExcel(deutschMeaning,path,sheetName =sheetName) is not False):
                deutschMeaningValSheetCoRow = self.findValueInExcel(deutschMeaning,path,sheetName =sheetName)
                # print(f"{deutschMeaningValSheetCoRow = }")
                row = deutschMeaningValSheetCoRow[0][2]
            elif ((deutschMeaning ==None and meaningId !=None) or (deutschMeaning !=None and meaningId !=None) ) and self.findValueInExcel(meaningId,path,sheetName =sheetName) is not False:
                meaningIdValSheetCoRow = self.findValueInExcel(meaningId,path,sheetName =sheetName)
                # print(f"{meaningIdValSheetCoRow = }")
                row = meaningIdValSheetCoRow[0][2]
            meaningIdIn,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer =[],None,None,None,None,None,None,None,None,None,None,None,None,None

            meaningIdTitleExSheetCoRow = self.findValueInExcel(self.meaningIdTitleEx,path,sheetName =sheetName)
            for x in range(len(meaningIdTitleExSheetCoRow)):
                meaningIdIn.append(mySheet.cell(row,meaningIdTitleExSheetCoRow[x][1]).value)

            
            mNumDafTitleExESheetCoRow = self.findValueInExcel(self.mNumDafTitleEx,path,sheetName =sheetName)
            meaningNumDAF = mySheet.cell(row,mNumDafTitleExESheetCoRow[0][1]).value

            deuTitleExESheetCoRow = self.findValueInExcel(self.deuTitleEx,path,sheetName =sheetName)
            deutsch = mySheet.cell(row,deuTitleExESheetCoRow[0][1]).value


            engTitleExSheetCoRow = self.findValueInExcel(self.engTitleEx,path,sheetName =sheetName)
            english = mySheet.cell(row,engTitleExSheetCoRow[0][1]).value

            synTitleExSheetCoRow = self.findValueInExcel(self.synTitleEx,path,sheetName =sheetName)
            synonym = mySheet.cell(row,synTitleExSheetCoRow[0][1]).value

            perTitleExSheetCoRow = self.findValueInExcel(self.perTitleEx,path,sheetName =sheetName)
            persian = mySheet.cell(row,perTitleExSheetCoRow[0][1]).value

            nDeFileAuTitleExSheetCoRow = self.findValueInExcel(self.nDeFileAuTitleEx,path,sheetName =sheetName)
            nameDeAudio = mySheet.cell(row,nDeFileAuTitleExSheetCoRow[0][1]).value


            addTelDeuTitleExSheetCoRow = self.findValueInExcel(self.addTelDeuTitleEx,path,sheetName =sheetName)
            addTelDeu = mySheet.cell(row,addTelDeuTitleExSheetCoRow[0][1]).value

            naSynFileAuExSheetCoRow = self.findValueInExcel(self.naSynFileAuEx,path,sheetName =sheetName)
            nameSynAudio = mySheet.cell(row,naSynFileAuExSheetCoRow[0][1]).value

            addTelSynTitleExSheetCoRow = self.findValueInExcel(self.addTelSynTitleEx,path,sheetName =sheetName)
            addTelSyn = mySheet.cell(row,addTelSynTitleExSheetCoRow[0][1]).value

            naEngFileAuTitleExSheetCoRow = self.findValueInExcel(self.naEngFileAuTitleEx,path,sheetName =sheetName)
            nameEnAudio = mySheet.cell(row,naEngFileAuTitleExSheetCoRow[0][1]).value

            addTelEngTitleExSheetCoRow = self.findValueInExcel(self.addTelEngTitleEx,path,sheetName =sheetName)
            addTelEn = mySheet.cell(row,addTelEngTitleExSheetCoRow[0][1]).value

            naPerFileAudioTitleExSheetCoRow = self.findValueInExcel(self.naPerFileAudioTitleEx,path,sheetName =sheetName)
            namePerAudio = mySheet.cell(row,naPerFileAudioTitleExSheetCoRow[0][1]).value

            addTelPerSheetCoRow = self.findValueInExcel(self.addTelPer,path,sheetName =sheetName)
            addTelPer = mySheet.cell(row,addTelPerSheetCoRow[0][1]).value
            output = [meaningIdIn,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer]
            # return meaningId,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer
        # else:
        return output


    def getFilledMeaningWordAllDetails(self,word,deutschMeaning=None,meaningId=None):
        mySheet = self.getSheet(self.path,self.sheetName)
        path = self.path
        sheetName = self.sheetName
        row = None
        if self.findValueInExcel(word,path,sheetName =sheetName):
            # row = ""
            # print(f"{row = }")
            if deutschMeaning ==None and meaningId==None or (self.findValueInExcel(deutschMeaning,path,sheetName =sheetName) is False) or (self.findValueInExcel(meaningId,path,sheetName =sheetName) is False):
                wordValSheetCoRow = self.findValueInExcel(word,path,sheetName =sheetName)
                # print(f"{wordValSheetCoRow = }")
                # print(f"{row = }")
                row = wordValSheetCoRow[0][2]


            elif deutschMeaning !=None and meaningId ==None and (self.findValueInExcel(deutschMeaning,path,sheetName =sheetName) is not False):
                deutschMeaningValSheetCoRow = self.findValueInExcel(deutschMeaning,path,sheetName =sheetName)
                # print(f"{deutschMeaningValSheetCoRow = }")
                row = deutschMeaningValSheetCoRow[0][2]
            elif ((deutschMeaning ==None and meaningId !=None) or (deutschMeaning !=None and meaningId !=None) ) and self.findValueInExcel(meaningId,path,sheetName =sheetName) is not False:
                meaningIdValSheetCoRow = self.findValueInExcel(meaningId,path,sheetName =sheetName)
                # print(f"{meaningIdValSheetCoRow = }")
                row = meaningIdValSheetCoRow[0][2]
            meaningIdIn,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer =None,None,None,None,None,None,None,None,None,None,None,None,None,None

            meaningIdTitleExSheetCoRow = self.findValueInExcel(self.meaningIdTitleEx,path,sheetName =sheetName)
            meaningIdIn = mySheet.cell(row,meaningIdTitleExSheetCoRow[0][1]).value

            
            mNumDafTitleExESheetCoRow = self.findValueInExcel(self.mNumDafTitleEx,path,sheetName =sheetName)
            meaningNumDAF = mySheet.cell(row,meaningIdTitleExSheetCoRow[0][1]).value

            deuTitleExESheetCoRow = self.findValueInExcel(self.deuTitleEx,path,sheetName =sheetName)
            deutsch = mySheet.cell(row,deuTitleExESheetCoRow[0][1]).value


            engTitleExSheetCoRow = self.findValueInExcel(self.engTitleEx,path,sheetName =sheetName)
            english = mySheet.cell(row,engTitleExSheetCoRow[0][1]).value

            synTitleExSheetCoRow = self.findValueInExcel(self.synTitleEx,path,sheetName =sheetName)
            synonym = mySheet.cell(row,synTitleExSheetCoRow[0][1]).value

            perTitleExSheetCoRow = self.findValueInExcel(self.perTitleEx,path,sheetName =sheetName)
            persian = mySheet.cell(row,perTitleExSheetCoRow[0][1]).value

            nDeFileAuTitleExSheetCoRow = self.findValueInExcel(self.nDeFileAuTitleEx,path,sheetName =sheetName)
            nameDeAudio = mySheet.cell(row,nDeFileAuTitleExSheetCoRow[0][1]).value


            addTelDeuTitleExSheetCoRow = self.findValueInExcel(self.addTelDeuTitleEx,path,sheetName =sheetName)
            addTelDeu = mySheet.cell(row,addTelDeuTitleExSheetCoRow[0][1]).value

            naSynFileAuExSheetCoRow = self.findValueInExcel(self.naSynFileAuEx,path,sheetName =sheetName)
            nameSynAudio = mySheet.cell(row,naSynFileAuExSheetCoRow[0][1]).value

            addTelSynTitleExSheetCoRow = self.findValueInExcel(self.addTelSynTitleEx,path,sheetName =sheetName)
            addTelSyn = mySheet.cell(row,addTelSynTitleExSheetCoRow[0][1]).value

            naEngFileAuTitleExSheetCoRow = self.findValueInExcel(self.naEngFileAuTitleEx,path,sheetName =sheetName)
            nameEnAudio = mySheet.cell(row,naEngFileAuTitleExSheetCoRow[0][1]).value

            addTelEngTitleExSheetCoRow = self.findValueInExcel(self.addTelEngTitleEx,path,sheetName =sheetName)
            addTelEn = mySheet.cell(row,addTelEngTitleExSheetCoRow[0][1]).value

            naPerFileAudioTitleExSheetCoRow = self.findValueInExcel(self.naPerFileAudioTitleEx,path,sheetName =sheetName)
            namePerAudio = mySheet.cell(row,naPerFileAudioTitleExSheetCoRow[0][1]).value

            addTelPerSheetCoRow = self.findValueInExcel(self.addTelPer,path,sheetName =sheetName)
            addTelPer = mySheet.cell(row,addTelPerSheetCoRow[0][1]).value
            verification = [meaningId,meaningNumDAF,deutsch,english,synonym,persian]
            if None in verification:
                return False
            else:
                return meaningId,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer
        else:
            return False

            # output= [bookName,chapter,content,page,publisher,meaningId] 
    def getBookAllcontentFromExcel(self,word):
        output = []
        mySheet = self.getSheet(self.path,self.sheetName)
        path = self.path
        sheetName = self.sheetName
        row = None
        if self.findValueInExcel(word,path,sheetName =sheetName):
            bookName,chapter,content,page,publisher,meaningId = None,None,None,None,None,None
            wordValSheetCoRow = self.findValueInExcel(word,path,sheetName =sheetName)
            row = wordValSheetCoRow[0][2]
            # self.chapTitleEx = "chapter"
            # self.contTitleEx = "content"
            # self.pageTitleEx = "page"
            # self.pubTitleEx = "publisher"
            # self.bookNTitleEx = "book name"

            chapTitleExSheetCoRow = self.findValueInExcel(self.chapTitleEx,path,sheetName =sheetName)
            chapter = mySheet.cell(row,chapTitleExSheetCoRow[0][1]).value

            contTitleExSheetCoRow = self.findValueInExcel(self.contTitleEx,path,sheetName =sheetName)
            content = mySheet.cell(row,contTitleExSheetCoRow[0][1]).value

            pageTitleExSheetCoRow = self.findValueInExcel(self.pageTitleEx,path,sheetName =sheetName)
            page = mySheet.cell(row,pageTitleExSheetCoRow[0][1]).value

            pubTitleExSheetCoRow = self.findValueInExcel(self.pubTitleEx,path,sheetName =sheetName)
            publisher = mySheet.cell(row,pubTitleExSheetCoRow[0][1]).value

            bookNTitleExSheetCoRow = self.findValueInExcel(self.bookNTitleEx,path,sheetName =sheetName)
            bookName = mySheet.cell(row,bookNTitleExSheetCoRow[0][1]).value

            meaningIditleExSheetCoRow = self.findValueInExcel(self.meaningIdTitleEx,path,sheetName =sheetName)
            meaningId = mySheet.cell(row,meaningIditleExSheetCoRow[0][1]).value

            output= [bookName,chapter,content,page,publisher,meaningId]      
        
        return output

    def getBookAllcontentFromMongo(self,word):
        output = []
        bookName,chapter,content,page,publisher,meaningId = None,None,None,None,None,None
        condition = {"word":word}
        outputMongo = {"bookName" : 1,"chapter":1,"content":1,"page":1,"publisher":1,"meaningId":1,"_id":0}
        for col in self.obj_Db.bookCollection.find(condition,output):
            for field,val in col.items():
                if field == "bookName" :
                   bookName = val
                   output.append(bookName)
                if field == "chapter":
                   chapter = val
                   output.append(chapter)
                if field == "content":
                   content = val
                   output.append(content)
                if field ==  "page":
                   page = val
                   output.append(page)
                if field == "publisher":
                   publisher = val
                   output.append(publisher)
                if field ==  "meaningId":
                   meaningId = val
                   output.append(meaningId)
        return output    

    def getFilledBookAllcontent(self,word):
            mySheet = self.getSheet(self.path,self.sheetName)
            path = self.path
            sheetName = self.sheetName
            row = None
            if self.findValueInExcel(word,path,sheetName =sheetName):
                bookName,chapter,content,page,publisher = None,None,None,None,None
                wordValSheetCoRow = self.findValueInExcel(word,path,sheetName =sheetName)
                row = wordValSheetCoRow[0][2]

                chapTitleExSheetCoRow = self.findValueInExcel(self.chapTitleEx,path,sheetName =sheetName)
                chapter = mySheet.cell(row,chapTitleExSheetCoRow[0][1]).value

                contTitleExSheetCoRow = self.findValueInExcel(self.contTitleEx,path,sheetName =sheetName)
                content = mySheet.cell(row,contTitleExSheetCoRow[0][1]).value

                pageTitleExSheetCoRow = self.findValueInExcel(self.pageTitleEx,path,sheetName =sheetName)
                page = mySheet.cell(row,pageTitleExSheetCoRow[0][1]).value

                pubTitleExSheetCoRow = self.findValueInExcel(self.pubTitleEx,path,sheetName =sheetName)
                publisher = mySheet.cell(row,pubTitleExSheetCoRow[0][1]).value

                bookNTitleExSheetCoRow = self.findValueInExcel(self.bookNTitleEx,path,sheetName =sheetName)
                bookName = mySheet.cell(row,bookNTitleExSheetCoRow[0][1]).value
                verificationFilled = [bookName,chapter,content,page,publisher]
                if None in verificationFilled:
                    return False
                else:
                    return bookName,chapter,content,page,publisher      
            else:
                return False  
    def getIdFromMongo(self,word):
        condition = {"word":word}
        outputMongo = {"_id":1}
        # word , book , meaning
        output = [None,None,None]
        for col in self.obj_Db.wordCollection.find(condition,outputMongo):
            for keys in col.keys():
                output[0] = col[keys]
        for col in self.obj_Db.bookCollection.find(condition,outputMongo):
            for keys in col.keys():
                output[1] = col[keys]
        for col in self.obj_Db.meaningCollection.find(condition,outputMongo):
            for keys in col.keys():
                output[2] = col[keys]
        if None in output:
            output = False
        return output
                         

class Entering:
    def __init__(self):
        self.obj_Db = Db()
        self.obj_Finding = Finding()
        self.backupAdd = r"D:\project\WoerterbuchProject\mainV2\dataEntry\wordsNew.xlsx"
         
    def enterMongoWordAllContents(self,word,meaningId,wordType=None,article=None,voiceName=None,voiceTelLink=None):
        output = False
        verficationExistense = False
        condition = {"word":word,"meaningId":{"$in":[ObjectId(meaningId)]}}
        outputFieldNVal = {"_id":1}
        for col in self.obj_Db.wordCollection.find(condition,outputFieldNVal):
            for keys in col.keys():
                # if word == col[keys]:
                verficationExistense = True
        wordDetails = {"word":word,"wordType":wordType,"article":article,"voice":{"name":voiceName,"telegramLink":voiceTelLink},"meaningId":[ObjectId(meaningId)]}
        # print(f"{verficationExistense = }")
        if verficationExistense is False:
            self.obj_Db.wordCollection.insert_one(wordDetails)
            wordId = ""
            # output = True
            for col in self.obj_Db.wordCollection.find(wordDetails,outputFieldNVal):
                for keys in col.keys():
                    output = col[keys]
            # findValueInExcel(self,val,pathFile,sheetName =None):
            path = self.obj_Finding.path
            sheetName = self.obj_Finding.sheetName
            wordIdColNRow = self.obj_Finding.findValueInExcel(self.obj_Finding.wordIdTitleEx,path,sheetName)
            #FIXME enterMongoWordAllContents update excel word id 
        return output 

                # 
   
    
    def enterMeaningWordAllContents(self,word,deutsch,english=None,synonym=None,persian=None,meaningNumDaf=None,deuName=None,deuTelegramLink=None,engName=None,engTelegramLink=None,synName=None,synTelegramLink=None,perName=None,perTelegramLink=None):
        output = False
        verficationExistense = False
        condition = {"deutsch":deutsch,"english":english,"synonym":synonym,"persian":persian,"meaningNumDaf":meaningNumDaf,"word":word}
        outputFieldNVal = {"_id":0,"word":1}
        for col in self.obj_Db.meaningCollection.find(condition,outputFieldNVal):
            for keys in col.keys():
                if word == col[keys]:
                    verficationExistense = True
        meaningDetails = {"deutsch":deutsch,"english":english,"synonym":synonym,"persian":persian,"meaningNumDaf":meaningNumDaf,"word":word,"deuVoice":{"deuName":deuName,"deuTelegramLink":deuTelegramLink},"engVoice":{"engName":engName,"engTelegramLink":engTelegramLink},"synVoice":{"synName":synName,"synTelegramLink":synTelegramLink},"perVoice":{"perName":perName,"perTelegramLink":perTelegramLink}}
        # print(f"{verficationExistense = }")
        if verficationExistense is False:
            self.obj_Db.meaningCollection.insert_one(meaningDetails)
            # print(f"in method ")
            outputFieldNVal2 = {"_id":1}
            for col in self.obj_Db.meaningCollection.find(condition,outputFieldNVal2):
                for keys in col.keys():
                    output = col[keys]
                    # print(f"from method {output = }")
                    # (condition,{"$set":meaningDetails})

            ## Enter meaning id in word collection process
            meaningIdInWordCollect = []
            for col in self.obj_Db.wordCollection.find({"word":word},{"meaningId":1,"_id":0}):
                for keys in col.keys():
                    meaningIdInWordCollect.append(col[keys])
            meaningIdList = [] 
            # print(f"{output = }")
            meaningIdList.append(output)
            # print(f"{meaningIdInWordCollect = }")
            meaningIds = meaningIdList + meaningIdInWordCollect
            # print(f"{meaningIds = }")
            updateVal = {"meaningId":meaningIds}
            # print(f"{updateVal = }")
            self.obj_Db.wordCollection.update_many({"word":word},{"$set":updateVal})
            ##
        return output 



    def enterMongoBookAllContents(self,word,meaningId,bookName=None,chapter=None,content=None,page=None,publisher=None):
        output = False
        wordExistence = None
        verificationWord = False
        condition = {"word":word,"meaningId":meaningId}
        for col in self.obj_Db.bookCollection.find(condition,{"_id":1}):
            for keys in col.keys():
                verificationWord = True
        if verificationWord is False:
            bookDetails = {"word":word,"bookName":bookName,"chapter":chapter,"content":content,"page":page,"publisher":publisher,"meaningId":meaningId}
            self.obj_Db.bookCollection.insert_one(bookDetails)
            outputMongo = {"_id":1}
            for col in self.obj_Db.bookCollection.find(bookDetails,outputMongo):
                for keys in col.keys():
                    output = col[keys]
        return output

class Updating:
    def __init__(self):
        self.obj_Finding = Finding()
        self.obj_Entering = Entering()
        self.backupPath = r"D:\project\WoerterbuchProject\mainV2\dataEntry\wordsBackup.xlsx"
        self.obj_Db = Db()
        
    def updateExcelId(self,mySheet,title,idValue,word,pathNName,sheetName):
        output = False
        column=None
        # rowOptional=None
        if title == self.obj_Finding.wordIdTitleEx:
            # pass
            sheetColRow = self.obj_Finding.findValueInExcel(title,pathNName,sheetName=sheetName)
            column = sheetColRow[0][1]
            # optionalSheetColRow = self.obj_Finding.findValueInExcel(address_Telegram_wort,pathNName,sheetName=sheetName)
            # rowOptional = optionalSheetColRow[0][2]
        elif title == self.obj_Finding.bookIdTitleEx:
            # pass
            sheetColRow = self.obj_Finding.findValueInExcel(title,pathNName,sheetName=sheetName)
            column = sheetColRow[0][1]
            # optionalSheetColRow = self.obj_Finding.findValueInExcel(page,pathNName,sheetName=sheetName)
            # rowOptional = optionalSheetColRow[0][2]
        elif title == self.obj_Finding.meaningIdTitleEx: 
            # pa/ss   
            sheetColRow = self.obj_Finding.findValueInExcel(title,pathNName,sheetName=sheetName)
            column = sheetColRow[0][1]
            # optionalSheetColRow = self.obj_Finding.findValueInExcel(deutsch,pathNName,sheetName=sheetName)
            # rowOptional = optionalSheetColRow[0][2]

        wordSheetColRow = self.obj_Finding.findValueInExcel(word,pathNName,sheetName=sheetName)
        row = wordSheetColRow[0][2]
        # wb = load_workbook(filename=pathNName)
        # mySheet = wb.get_sheet_by_name(sheetName)
        if mySheet.cell(row,column).value !=idValue:
            output = True
            # mySheet.cell(row,column).value = str(idValue)

            # wb.save(pathNName)
            # wb.save(self.backupPath)
        return output

    #def  updateWordDetailsMongo(self,wordId,word=word,meaningId=meaningId,wordType=wordType,article=article,voiceName=voiceName,voiceTelLink=voiceTelLink):
    def updateWordDetailsMongo(self,wordId,word=None,meaningId=None,wordType=None,article=None,voiceName=None,voiceTelLink=None):
        check = [word,meaningId,wordType,article,voiceName,voiceTelLink]
        output = check.copy()

        upWord,upMeaningId,upWordType,upArticle,upVoiceName,upVoiceTelLink=None,None,None,None,None,None
        changedWordId = None
        if isinstance(wordId,str):
            changedWordId = ObjectId(wordId)
        else:
            changedWordId = wordId
        condition = {"_id":changedWordId}
        temp = None
        outputMongo = {"word":1,"meaningId":1,"wordType":1,"article":1,"voice":1,"_id":0}
        for col in self.obj_Db.wordCollection.find(condition,outputMongo):
            for field,val in col.items():
                # print(f"{field = }\n{val = }")
                if field == "meaningId":
                   upMeaningId = val
                if field == "wordType":
                   upWordType = val
                if field == "article":
                   upArticle = val
                if field == "voice":
                   print(f"voice for loop {val = }")
                   upVoiceName = val["name"]
                   upVoiceTelLink = val["telegramLink"]
        
        # for x in range(len(check)):
        if check[0] !=None and check[0]!=upWord:
            # pass
            output[0]=False
        if check[1] !=None and (check[1]!=upMeaningId or check[1] not in upMeaningId):
            changed = None
            if isinstance(upMeaningId,list):
                withoutRepeatedUpMeaningId = list(set(upMeaningId))
                changed =  withoutRepeatedUpMeaningId + [ObjectId(check[1])]
            else:
                changed =  [ObjectId(check[1])]
            self.obj_Db.wordCollection.update_many(condition,{"$set":{"meaningId":changed}})
            output[1]=True
        if check[2] != None and check[2]!=upWordType:
            self.obj_Db.wordCollection.update_many(condition,{"$set":{"wordType":check[2]}})
            output[2]=True
        if check[3] != None and check[3]!=upArticle:
            self.obj_Db.wordCollection.update_many(condition,{"$set":{"article":check[3]}})
            output[3]=True
        voiceVrification = False
        voiceUpdate = {"voice":{"name":upVoiceName,"telegramLink":upVoiceTelLink}}
        if check[4] != None and check[4]!=upVoiceName:
            voiceVrification = True
            voiceUpdate["voice"]["name"]=check[4]
            # print(f"{check[4] = }\n{upVoiceName = }\n{voiceUpdate = }")
            output[4]=True 
            
        if check[5] != None and check[5]!=upVoiceTelLink:
            voiceVrification = True
            voiceUpdate["voice"]["telegramLink"]=check[5]
            # print(f"{check[5] = }\n{upVoiceTelLink =}\n{voiceUpdate = }")
            output[5]=True
        if voiceVrification is True:
            self.obj_Db.wordCollection.update_many(condition,{"$set":voiceUpdate})
            # print(f"update voice {check[4] = }\n{check[5] = }\n{voiceUpdate = }") 
        return output




    def updateMeaningDetailsMongo(self,meaningId,word=None,wordType=None,article=None,deutsch=None,english=None,synonym=None,persian=None,meaningNumDaf=None,deuName=None,deuTelegramLink=None,engName=None,engTelegramLink=None,synName=None,synTelegramLink=None,perName=None,perTelegramLink=None):
        # (meaningId,word=wordDetails[0],wordType=wordDetails[1],article=wordDetails[2],deutsch=deutsch,english=english,synonym=synonym,persian=persian,meaningNumDaf=meaningNumDaf,deuName=voice[0],deuTelegramLink=voice[1],engName=voice[2],engTelegramLink=voice[3],synName=voice[4],synTelegramLink=voice[5],perName=voice[6],perTelegramLink=voice[7])
        changedMeaningId = None
        if isinstance(meaningId,str):
            changedMeaningId = ObjectId(meaningId)
        else:
            changedMeaningId = meaningId

        output = [False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False]
        inputCheck = [word,wordType,article,deutsch,english,synonym,persian,meaningNumDaf,deuName,deuTelegramLink,engName,engTelegramLink,synName,synTelegramLink,perName,perTelegramLink]
        testCheck = [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None]
        wordDetails,deuVoice,engVoice,synVoice,perVoice=None,None,None,None,None
        fields = ["word","wordType","article","deutsch","english","synonym","persian","meaningNumDaf","deuName","deuTelegramLink","engName","engTelegramLink","synName","synTelegramLink","perName","perTelegramLink"]
        condition = {"_id":changedMeaningId}
        ouputMongo = {"wordDetails":1,"deutsch":1,"english":1,"synonym":1,"persian":1,"meaningNumDaf":1,"deuVoice":1,"engVoice":1,"synVoice":1,"perVoice":1,"_id":0}
        verificationId = False
        for col in self.obj_Db.meaningCollection.find(condition,ouputMongo):
           for field,val in col.items():
                verificationId = True
                if field == "wordDetails":
                   testCheck[0]=val[fields[0]]
                   testCheck[1]=val[fields[1]]
                   testCheck[2]=val[fields[2]]
                   wordDetails = val
                if field == fields[3]:
                   testCheck[3]=val 
                if field == fields[4]:
                   testCheck[4]=val
                if field == fields[5]:
                   testCheck[5]=val 
                if field == fields[6]:
                   testCheck[6]=val
                if field == fields[7]:
                   testCheck[7]=val
                if field == "deuVoice":
                   testCheck[8]=val[fields[8]]
                   testCheck[9]=val[fields[9]] 
                   deuVoice = val
                if field == "engVoice":
                   testCheck[10]=val[fields[10]] 
                   testCheck[11]=val[fields[11]]
                   engVoice = val
                if field == "synVoice":
                   testCheck[12]=val[fields[12]] 
                   testCheck[13]=val[fields[13]] 
                   synVoice = val
                if field == "perVoice":
                   testCheck[14]=val[fields[14]] 
                   testCheck[15]=val[fields[15]]
                   perVoice = val

        if verificationId is True:
            for x in range(len(inputCheck)):
                if inputCheck[x] != None and inputCheck[x]!=testCheck[x]:
                    valUpdate = None
                    if x == 0 or x == 1 or x ==2 :
                       wordDetails.update({fields[x]:inputCheck[x]}) 
                       valUpdate = {"wordDetails":wordDetails}
                    elif x==8 or x==9:
                       deuVoice.update({fields[x]:inputCheck[x]}) 
                       valUpdate = {"deuVoice":deuVoice}
                    elif x==10 or x==11:
                       engVoice.update({fields[x]:inputCheck[x]}) 
                       valUpdate = {"engVoice":engVoice}
                    elif x==12 or x==13:
                       synVoice.update({fields[x]:inputCheck[x]}) 
                       valUpdate = {"synVoice":synVoice}
                    elif x==14 or x==15:
                       perVoice.update({fields[x]:inputCheck[x]}) 
                       valUpdate = {"perVoice":perVoice}
                    else:
                        valUpdate={fields[x]:inputCheck[x]}

                    self.obj_Db.meaningCollection.update_many(condition,{"$set":valUpdate})
                    output[x]=True
        return output








        # output=False
        # verficationExistense = False
        # # condition = {"deutsch":deutsch,"english":english,"synonym":synonym,"persian":persian,"meaningNumDaf":meaningNumDaf,"word":word}
        # condition = {"word":word}

        # outputFieldNVal = {"_id":0,"word":1}
        # for col in self.obj_Db.meaningCollection.find(condition,outputFieldNVal):
        #     # print(f"{col = }")
        #     for keys in col.keys():
        #         # if word == col[keys]:
        #             verficationExistense = True
        # meaningDetails = {"deutsch":deutsch,"english":english,"synonym":synonym,"persian":persian,"meaningNumDaf":meaningNumDaf,"word":word,"deuVoice":{"deuName":deuName,"deuTelegramLink":deuTelegramLink},"engVoice":{"engName":engName,"engTelegramLink":engTelegramLink},"synVoice":{"synName":synName,"synTelegramLink":synTelegramLink},"perVoice":{"perName":perName,"perTelegramLink":perTelegramLink}}
        # # # print(f"{verficationExistense = }")
        # if verficationExistense is True:
        #         self.obj_Db.meaningCollection.update_many(condition,{"$set":meaningDetails})
        #         output= True

        return output 

    def updateBookDetailsMongo(self,bookId,word=None,bookName=None,chapter=None,content=None,page=None,publisher=None,meaningId=None):
        inputCheck = [word,bookName,chapter,content,page,publisher,meaningId]
        output = [False,False,False,False,False,False,False]
        opWord,opBookName,opChapter,opContent,opPage,opPublisher,opMeaningId=None,None,None,None,None,None,None
        getCheck=[None,None,None,None,None,None,None]
        fields = ["word","bookName","chapter","content","page","publisher","meaningId"]
        changedBookId = None
        if isinstance(bookId,str):
            changedBookId = ObjectId(bookId)
        else:
            changedBookId = bookId
        condition = {"_id":changedBookId}
        outputMongo = {"word":1,"bookName":1,"chapter":1,"content":1,"page":1,"publisher":1,"meaningId":1,"_id":0}
        verificationExistence = False
        for col in self.obj_Db.bookCollection.find(condition,outputMongo):
            for field,val in col.items():
                verificationExistence = True
                if field == "word":
                    # opWord = val
                    getCheck[0]=val
                if field == "bookName":
                    # opBookName = val 
                    getCheck[1]=val
                if field == "chapter":
                    # opChapter = val
                    getCheck[2]=val
                if field == "content":
                    # opContent = val
                    getCheck[3]=val
                if field == "page":
                    # opPage = val
                    getCheck[4]=val 
                if field == "publisher":
                    # opPublisher = val
                    getCheck[5]=val
                if field == "meaningId":
                    # opMeaningId = val
                    getCheck[6]=val
        # opCheck = []
        # print(f"{verificationExistence = }")
        if verificationExistence is True:
            for x in range(len(getCheck)):
                if x != 0 and getCheck[x] != inputCheck[x] and inputCheck[x]!=None:
                    updatedFieldNVal = {fields[x] :inputCheck[x]}
                    self.obj_Db.bookCollection.update_many(condition,{"$set":updatedFieldNVal})
                    output[x]=True
        return output

                    



        verification = False
        condition = {"word":word,"meaningId":meaningId}
        # condition = {"word":word}
        for col in self.obj_Db.bookCollection.find(condition,{"_id":1}):
            for keys in col.keys():
                 verification = True
        if verification is True:   
            updateFilNVal = {"word":word,"bookName":bookName,"chapter":chapter,"content":content,"page":page,"publisher":publisher,"meaningId":meaningId}
            self.obj_Db.bookCollection.update_many(condition,{"$set":updateFilNVal})
            output = True
        return output

    def updateOneIdsWordBookMeaningExcel(self,mySheet,word,idsList,pathNName,sheetName):
        output = False
        # ids variable wordIdTitleEx bookIdTitleEx meaningIdTitleEx
        # updateExcelId(title,idValue,word,pathNName,sheetName)
        if idsList is not False:
            self.updateExcelId(mySheet,self.obj_Finding.wordIdTitleEx,idsList[0],word,pathNName,sheetName)
            self.updateExcelId(mySheet,self.obj_Finding.bookIdTitleEx,idsList[1],word,pathNName,sheetName)
            self.updateExcelId(mySheet,self.obj_Finding.meaningIdTitleEx,idsList[2],word,pathNName,sheetName)
            output = True
        return output

    def updateAllIdsWordBookMeaningExcel(self,pathNName,sheetName):
        output = []
        _,words = self.obj_Finding.getExcelWords(pathNName,sheetName)
        # words = []
        # for each in wordsNumTuple:
        #     words.append(each[1])

        mySheet,wb = self.obj_Finding.getSheetNWorkbook(pathNName,sheetName)
        # print(f"{words = }")
        for i in words:
            idsList = self.obj_Finding.getIdFromMongo(i)
            temp = []
            if self.updateOneIdsWordBookMeaningExcel(mySheet,i,idsList,pathNName,sheetName) is not False:
                temp.append(i)
                temp.append(True)
            else:
                temp.append(i)
                temp.append(False)
            output.append(tuple(temp))
        wb.save(pathNName)
        wb.save(self.backupPath)
        return output
class Execute:
    def __init__(self):
        self.obj_finding = Finding()
        self.obj_Db = Db()
        self.obj_Updating = Updating()
        self.obj_Entering = Entering()
        pass

    def executeUpdate(self):
        path = self.obj_finding.path
        sheetName = self.obj_finding.sheetName
        mySheet = self.obj_finding.getSheet(path,sheetName)
        _,words = self.obj_finding.getExcelWords(path,sheetName)
        idSheetColNRow = self.obj_finding.findValueInExcel(self.obj_finding.meaningIdTitleEx,path,sheetName)
        columnId = idSheetColNRow[0][1]
        output = [False,[],False]
        for eachWord in words:
            
            wordSheetColNRow = self.obj_finding.findValueInExcel(eachWord,path,sheetName)    
            #  word contenst update part
            wordType,article,voiceName,voiceTelLink,meaningId = self.obj_finding.getWordsFromExcelAllContent(eachWord)
            outputWord = self.obj_Updating.updateWordDetailsMongo(eachWord,wordType,article,meaningId,voiceName=voiceName,voiceTelLink=voiceTelLink)
            if outputWord:
                output[0]=True

            # meaning contenst update part 
            outputIndex1 = []
            for x in range(len(wordSheetColNRow)):
                row = wordSheetColNRow[x][2]
                idVal = mySheet.cell(row,columnId).value
                outputMeaningEx = self.obj_finding.getMeaningWordFromExcelAllDetails(eachWord,meaningId=idVal)

                meaningId,meaningNumDaf,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer = outputMeaningEx[0],outputMeaningEx[1],outputMeaningEx[2],outputMeaningEx[3],outputMeaningEx[4],outputMeaningEx[5],outputMeaningEx[6],outputMeaningEx[7],outputMeaningEx[8],outputMeaningEx[9],outputMeaningEx[10],outputMeaningEx[11],outputMeaningEx[12],outputMeaningEx[13]

                outputMeaningMongo = self.obj_Updating.updateMeaningDetailsMongo(eachWord,deutsch,english,synonym,persian,meaningNumDaf,deuName=nameDeAudio,deuTelegramLink=addTelDeu,engName=nameEnAudio,engTelegramLink=addTelEn,synName=nameSynAudio,synTelegramLink=addTelSyn,perName=namePerAudio,perTelegramLink=addTelPer)
                outputIndex1.append(outputMeaningMongo)
            output[1]= tuple(outputIndex1)


            #book contenst update part 
            
            bookVerification = False
            outputIndex2 = []
            for y in range(len(wordSheetColNRow)):
                row = wordSheetColNRow[y][2]
                idVal = mySheet.cell(row,columnId).value
                outputBookEx =self.obj_finding.getBookAllcontentFromExcel(eachWord)
                # print(f"{outputBookEx = }")
                bookName,chapter,content,page,publisher,meaningId = outputBookEx[0],outputBookEx[1],outputBookEx[2],outputBookEx[3],outputBookEx[4],outputBookEx[5]
                outputBook = self.obj_Updating.updateBookDetailsMongo(eachWord,bookName,chapter,content,page,publisher,idVal)
                if outputBook is True:
                    outputIndex2.append(True)
                else:
                    outputIndex2.append(False)
            output[2]=tuple(outputIndex2)       

        return output
        
    def executeEnter(self,word,deutsch,english=None,synonym=None,persian=None,meaningNumDaf=None,deuName=None,deuTelegramLink=None,engName=None,engTelegramLink=None,synName=None,synTelegramLink=None,perName=None,perTelegramLink=None,wordType=None,article=None,voiceName=None,voiceTelLink=None,bookName=None,chapter=None,content=None,page=None,publisher=None):
        output = [False,False,False]
        meaningId = None
        outputMeaning = self.obj_Entering.enterMeaningWordAllContents(word,deutsch,english=english,synonym=synonym,persian=persian,meaningNumDaf=meaningNumDaf,deuName=deuName,deuTelegramLink=deuTelegramLink,engName=engName,engTelegramLink=engTelegramLink,synName=synName,synTelegramLink=synTelegramLink,perName=perName,perTelegramLink=perTelegramLink)
        if outputMeaning is not False:
            output[0]=outputMeaning
            meaningId= outputMeaning
        else:
            conditionMeaning = {"word":word,"deutsch":deutsch}
            for col in self.obj_Db.meaningCollection.find(conditionMeaning,{"_id":1}):
                for keys in col.keys():
                    # meaningVerfication = True
                    meaningId = col[keys]
        outputWord = self.obj_Entering.enterMongoWordAllContents(word,meaningId,wordType=wordType,article=article,voiceName=voiceName,voiceTelLink=voiceTelLink)
        if outputWord is not False:
            output[1] = outputWord

        outputBook = self.obj_Entering.enterMongoBookAllContents(word,meaningId,bookName=bookName,chapter=chapter,content=content,page=page,publisher=publisher)
        if outputBook is not False:
            output[2] = outputBook

        return output
    def executeIdsToExcelUpdate(self):
        pass

class Adding:
    def __init__(self):
        self.obj_Db = Db()
    def addWordDetailsToMeaningFromMongoToMongo(self):
        output = []
        # for col in self.obj_Db.meaningCollection.find({},{"_id":1}):
        #     for keys in col.keys():
        #         for x in col[keys]:
        outputMongo = {"_id":1,"word":1,"wordType":1,"article":1,"meaningId":1}
        for col in self.obj_Db.wordCollection.find({},outputMongo):
            wordId,word,wordType,article,meaningIdList=None,None,None,None,None
            temp=[]
            for field,val in col.items():
                if field == "_id":
                    wordId = val
                    temp.append(wordId)
                elif field == "word":
                    word = val
                    temp.append(word)
                elif field == "wordType":
                    wordType = val
                    temp.append(wordType)
                elif field == "article":
                    article = val
                    temp.append(article)
                elif field == "meaningId":
                    meaningIdList = val
            output.append(tuple(temp))        
            updatedVal={"wordDetails":{"wordId":wordId,"word":word,"wordType":wordType,"article":article}}
            for x in meaningIdList:
                self.obj_Db.meaningCollection.update_many({"_id":x},{"$set":updatedVal})

        return output
                


        pass

    

    def addMeaningIdToWordCollecFromMongoToMongo(self):
        #meathod 1
        output = []
        meaningIdNWords = []
        for col in self.obj_Db.meaningCollection.find({},{"_id":1,"word":1}):
            temp = []
            for field, val in col.items():
                
                if field == "_id":
                    temp.append(val)
                elif field == "word":
                    temp.append(val)
            meaningIdNWords.append(tuple(temp))
        wordIdsNWords=[]        
        for i in meaningIdNWords:
            for j in range(len(i)):
                if j==1:
                    for col in self.obj_Db.wordCollection.find({"word":i[j]},{"_id":1,"word":1,"meaningId":1}):
                        temp = []    
                        for field,val in col.items():
                            
                            if field == "_id":
                                temp.append(val)
                            elif field == "word":
                                temp.append(val)    
                            elif field == "meaningId":
                                if isinstance(val,list):
                                    k = [valIn for valIn in val if valIn !=""]                                    
                                elif isinstance(val,str):
                                    if val == "":
                                        k = []

                                    else:
                                        k = [ObjectId(val)]
                                temp.append(k)
                        
                        wordIdsNWords.append(tuple(temp))
        # print(f"wordIdsNWords = {wordIdsNWords}")  
        # print(f"\n\n")
        # print(f"meaningIdNWords = {meaningIdNWords}")                   
        for i in wordIdsNWords:
            for j in meaningIdNWords:
                # print(f"i[1] = {i[1]}")
                # print(f"\n\n")
                # print(f"j = {j}")
                # print(f"j[1] = {j[1]}")
                if i[1]==j[1]:
                    temp = []
                    temp.append(i[1])
                    # print(f"i = {i}\nj = {j}")
                    tempUp = i[2]+[j[0]]
                    tempUWithoutRepeated = list(set(tempUp))
                    temp.append(tempUWithoutRepeated)
                    self.obj_Db.wordCollection.update_many({"_id":i[0]},{"$set":{"meaningId":tempUWithoutRepeated}})
                    output.append(tuple(temp))

        return output


                                
class Deleting:
    def __init__(self):
        self.obj_Db=Db()
    def deleteWord(self,wordId):
        output=False
        varification = False
        condition = {"_id":ObjectId(wordId)}
        for col in self.obj_Db.wordCollection.find(condition,{"word":1}):
            for keys in col.keys():
                verification = True
        if verification is True:
            self.obj_Db.wordCollection.delete_one(condition)
            output = True
        return output



        



class Editing:
    def __init__(self):
        self.obj_Db = Db()


    def editWord(self,wordId,word=None,wordType=None,article=None,voiceName=None,telegramLink=None):
        condition = {"_id":ObjectId(wordId)}

        if wordType!=None:
            valUpdate={"wordType":wordType}
        if article!=None:
            valUpdate={"article":article}
        if voiceName!=None:
            valUpdate={"voice.name":voiceName}
        if telegramLink!=None:
            valUpdate={"voice.telegramLink":telegramLink}
        verification = False
        for col in self.obj_Db.wordCollection.find(condition,{"_id":1}):
            for keys in col.keys():
                verification = True
        if verification is True:
            if word!=None:
                valUpdate={"word":word}
                self.obj_Db.wordCollection.update_many(condition,{"$set":valUpdate})
            if wordType!=None:
                valUpdate={"wordType":wordType}
                self.obj_Db.wordCollection.update_many(condition,{"$set":valUpdate})
            if article!=None:
                valUpdate={"article":article}
                self.obj_Db.wordCollection.update_many(condition,{"$set":valUpdate})
            if voiceName!=None:
                valUpdate={"voice.name":voiceName}
                self.obj_Db.wordCollection.update_many(condition,{"$set":valUpdate})
            if telegramLink!=None:
                valUpdate={"voice.telegramLink":telegramLink}
                self.obj_Db.wordCollection.update_many(condition,{"$set":valUpdate})



    def editingWordColField(self,newField,beforeField):
        if newField==beforeField:
            return False

        output = False
        condition = {beforeField:{"$exists":True}}
        for col in self.obj_Db.wordCollection.find(condition,{"id":1}):
            for field,val in col.items():
                if field == beforeField:
                    output = True
        
        changed = {"$rename":{beforeField:newField}}
        self.obj_Db.wordCollection.update_many(condition,changed)
        return output     
       

class Search:
    def __init__(self):
        self.obj_Db = Db()
        
    def simpleSearch(self,word,deutsch=None):
        output = False

        # wordOutput = {"wordId":None,"word":None,"wordType":None,"article":None,"wordVoiceName":None,"wordTelAdd":None}
        # meaningNbookOutput = {"meaningId":None,"meaningNumDaf":None,"deutschMeaning":None,"deuName":None,"deuTelegramLink":None,"englishMeaning":None,"engName":None,"engTelegramLink":None,"synonymMeaning":None,"synName":None,"synTelegramLink":None,"persianMeaning":None,"perName":None,"perTelegramLink":None,"bookId":None,"bookName":None,"chapter":None,"content":None,"page":None,"lesson":None,"publisher":None}
        verification = False
        for col in self.obj_Db.wordCollection.find({"word":word},{"_id"}):
            for keys in col.keys():
                verification = True
        if verification is True:
            output = []
            bookNum = 0
            # bookNumRow,meaningNumRow,wordNumRow
            
            if deutsch == None:
                meaningCondition= {"wordDetails.word":word}
            else :
                meaningCondition= {"wordDetails.word":word,"deutsch":deutsch}
            wordOutputMongo = {"_id":1,"word":1,"wordType":1,"article":1,"voice":1}
            for col in self.obj_Db.wordCollection.find({"word":word},wordOutputMongo):
                wordDic = {}
                oneWord = []
                for field,val in col.items():
                    if field == "_id":
                        wordDic["wordId"]=val
                    elif field == "word":
                        wordDic["word"]=val
                    elif field == "wordType":
                        wordDic["wordType"]=val
                    elif field == "article":
                        wordDic["article"]=val
                    elif field == "voice":
                        wordDic["wordVoiceName"]=val["name"]
                        wordDic["wordtelegramLink"]=val["telegramLink"]
                # oneWord.append(wordDic)
                
                output.append(wordDic)








            meaningOutputMongo = {"_id":1,"deutsch":1,"english":1,"synonym":1,"persian":1,"meaningNumDaf":1,"synVoice":1,"engVoice":1,"perVoice":1,"deuVoice":1}
            
            for x in range(len(output)):
                i=0
                for col in self.obj_Db.meaningCollection.find(meaningCondition,meaningOutputMongo):
                    meaningOutput = {}
                    i=i+1
                    #FIXME meaning10
                    for field,val in col.items():
                        if field == "_id":
                            meaningOutput["meaningId"]=val
                        elif field == "deutsch":
                            meaningOutput["deutschMeaning"]=val
                        elif field =="english":
                            meaningOutput["englishMeaning"]=val
                        elif field =="synonym":
                            meaningOutput["synonymMeaning"]=val
                        elif field =="persian":
                            meaningOutput["persianMeaning"]=val
                        elif field =="meaningNumDaf":
                            meaningOutput["meaningNumDaf"]=val
                        elif field =="synVoice":                    
                            meaningOutput["synName"]=val["synName"]
                            meaningOutput["synTelegramLink"]=val["synTelegramLink"]
                        elif field =="engVoice":
                            meaningOutput["engName"]=val["engName"]
                            meaningOutput["engTelegramLink"]=val["engTelegramLink"]
                        elif field =="perVoice":
                            meaningOutput["perName"]=val["perName"]
                            meaningOutput["perTelegramLink"]=val["perTelegramLink"]
                        elif field =="deuVoice":
                            meaningOutput["deuName"]=val["deuName"]
                            meaningOutput["deuTelegramLink"]=val["deuTelegramLink"]
                    print(f"meaning{i}")
                        # temp[f"meaning{i}"]=meaningOutput
                        
                    output[x][f"meaning{i}"]= meaningOutput

                for x in range(len(output)): 
                    for field,val in output[x].items():
                        for i in range(10):
                            if field==f"meaning{i}":
                                tempMeaning = val
                                bookCondition = {"word":word,"meaningId":val["meaningId"]} 
                                bookOutputMongo = {"_id":1,"bookName":1,"chapter":1,"content":1,"page":1,"lesson":1,"publisher":1}
                                k = 0
                                for col in self.obj_Db.bookCollection.find(bookCondition,bookOutputMongo):
                                    bookOutput = {}
                                    k=k+1
                                    for fieldBook , valBook in col.items():
                                        if fieldBook == "_id":
                                            bookOutput["bookId"]=valBook
                                        elif fieldBook == "bookName":
                                            bookOutput["bookName"]=valBook
                                        elif fieldBook == "chapter":
                                            bookOutput["chapter"]=valBook
                                        elif fieldBook == "bookName":
                                            bookOutput["bookName"]=valBook
                                        elif fieldBook == "content":
                                            bookOutput["content"]=valBook
                                        elif fieldBook == "page":
                                            bookOutput["page"]=valBook
                                        elif fieldBook == "lesson":
                                            bookOutput["lesson"]=valBook
                                        elif fieldBook == "publisher":
                                            bookOutput["publisher"]=valBook
                                    
                                    tempMeaning[f"book{k}"]= bookOutput  
                                
                                output[x][f"meaning{i}"]=tempMeaning


        return output


@pytest.fixture
def deleting():
    deleting=Deleting()
    return deleting       


@pytest.fixture
def search():
    search = Search()
    return search

@pytest.fixture
def editing():
    editing = Editing()
    return editing

@pytest.fixture
def adding():
    adding=Adding()
    return adding


@pytest.fixture
def execute():
    execute=Execute()
    return execute

@pytest.fixture
def db():
    db = Db()
    return db

@pytest.fixture
def finding():
    finding = Finding()
    return finding
@pytest.fixture
def entering():
    entering = Entering()
    return  entering   
@pytest.fixture
def updating():
    updating = Updating()
    return updating
@pytest.fixture
def filePathNSheet():
    pathNName = r"D:\project\WoerterbuchProject\mainV2\dataEntry\words.xlsx" 
    sheetName = "orginal"
    yield pathNName,sheetName

@pytest.fixture
def enterAWord(db):
    word = "test1"
    wordType = "wordType1"
    article = "article1"
    voiceName = "voiceName1"
    telegramLink = "telegramLink1"
    meaningId = ["meaningId1"]
    insVal = {"word":word,"wordType":wordType,"article":article,"voice":{"name":voiceName,"telegramLink":telegramLink},"meaningId":meaningId}
    wordId=None
    db.wordCollection.insert_one(insVal)
    for col in db.wordCollection.find(insVal,{"_id":1}):
        for keys in col.keys():
            wordId = col[keys]
    return wordId,word,wordType,article,voiceName,telegramLink,meaningId   
# enterAWord

class Testdb:
    def test_WordInMeaningNWordComparison(self,db):
        listMeaning = []
        listWord = []
        for colM in db.wordCollection.find({},{"word":1,"_id":0}):
            for keys in colM.keys():
                listWord.append(colM[keys])
        for colW in db.meaningCollection.find({},{"word":1,"_id":0}):
            for keys in colW.keys():
                listMeaning.append(colW[keys])
        if len(listMeaning) != len(listWord):
            difference = set(listWord)-set(listMeaning)
            # print(f"{difference = }")
        assert len(listMeaning) == len(listWord)        


class TestFinding:
    #FIXME test_findValueInExcelOptional
    def test_findValueInExcelOptional(self,finding,filePathNSheet,kwargs):
        pathNName,sheetName=filePathNSheet
        mySheet = finding.getSheet(pathNName,sheetName)
        output = finding.findValueInExcelOptional(pathNName,sheetName,**kwargs)
        word = ""
        #mainRows = [(row,column),....]
        mainRows = []
        wordColumn = ""
        lengthKwargs = len(kwargs) 
        lengthTest = 0
        if output is True:
            for field,value in kwargs:
                #FIXME test_findValueInExcelOptional
                if field == finding.wordTitleEx:
                    word = value
            for row in range(1,mySheet.max_row+1):
                # print(f"{row = }")
                for col in range(1,mySheet.max_column+1):
                    # print(f"{col = }")
                    tempList = []
                    cell_value = mySheet.cell(row,col).value
                    if cell_value == word:
                        wordColumn = col
                        mainRows.append(row)
            copyMainRows = mainRows.copy()
            for field,value in kwargs:
                for row in range(1,mySheet.max_row+1):
                    # print(f"{row = }")
                    for col in range(1,mySheet.max_column+1):
                        # print(f"{col = }")
                        tempList = []
                        cell_value = mySheet.cell(row,col).value
                        if cell_value == field and cell_value!=word:
                            for eachRow in copyMainRows:
                                temp = mySheet.cell(eachRow,col).value
                                if temp != value:
                                    mainRows.remove(eachRow)
    @pytest.mark.parametrize("word,deutsch",[
        ("tschau","tschau! Ciao"),
        ("effandi","effandi"),
        ("auf wiedersehen","2 auf Wiedersehen! verwendet, um sich von jemandem zu verabschieden")
    ])
    def test_findWordNDeutschInExcel(self,finding,filePathNSheet,word,deutsch):
        path,sheetName=filePathNSheet
        mySheet = finding.getSheet(path,sheetName)
        # output = [(row,column),(row,column)]    #True
        # output = [False]   #False
        output = finding.findWordNDeutschInExcel(word,deutsch,path,sheetName)
        wordTitle = finding.wordTitleEx
        deutschTitle = finding.deuTitleEx
        wordColumn,deutschColumn,deutschRow,wordRow,outputResult = None,None,[],[],[]
        assert isinstance(output,list)
        for x in output:
            assert isinstance(x,bool)or isinstance(x,tuple)
            if isinstance(x,tuple):
                assert len(x)==2
        # mainRow = 
        if output[0] is not False:
            for row in range(1,mySheet.max_row+1):
                for col in range(1,mySheet.max_column+1):
                    cell_value = mySheet.cell(row,col).value
                    if wordTitle == cell_value:
                        wordColumn = col
                    elif deutschTitle == cell_value:
                        deutschColumn = col
                    elif deutsch == cell_value:
                        deutschRow.append(row)
                    elif word == cell_value:
                        wordRow.append(row)
            intersection = set(deutschRow)&set(wordRow)
            for eachRow in intersection:
                temp = []
                temp.append(eachRow)
                temp.append(wordColumn)
                outputResult.append(tuple(temp))
            assert output == outputResult

            

                    
    def test_redundencyExcelDeutschValue(self):
        pass



    def test_getSheet(self,finding,filePathNSheet):
        pathNName,sheetName=filePathNSheet
        mysheet = finding.getSheet(pathNName,sheetName)
        assert isinstance(mysheet,openpyxl.worksheet.worksheet.Worksheet)

    def test_getWordsMongo(self,finding,db):
        output = finding.getWordsMongo()
        assert isinstance(output,list)
        for eachone in output:
            isinstance(eachone,str)
        test_words = []
        for col in db.wordCollection.find({},{"word":1,"_id":0}):
            for keys in col.keys():
                test_words.append(col[keys])
        assert len(output) == len(test_words)
        assert output == test_words

    def test_getWordsNWordIdsNBookIdsNmeaningIds(self,filePathNSheet,finding):
        pathNName,sheetName=filePathNSheet
        _,words = finding.getExcelWords(pathNName,sheetName)
        output = finding.getWordsNWordIdsNBookIdsNmeaningIds(pathNName,sheetName)
        assert isinstance(output,list)
        assert len(words) == len(output)
        for x in range(len(output)):
            assert isinstance(output[x],tuple)
            assert len(output[x])==4
            assert isinstance(output[x][0],str)
            assert isinstance(output[x][1],str)
            assert isinstance(output[x][2],str)
            assert isinstance(output[x][3],str)





    def test_orginalSheetExistance(self,finding,filePathNSheet):
        pathNName,sheetName=filePathNSheet
        mysheet = finding.getSheet(pathNName,sheetName)
        # wordNumTitleEx  wordIdTitleEx
        sheetColRow = finding.findValueInExcel(finding.wordIdTitleEx,pathNName,sheetName=sheetName)
        column = sheetColRow[0][1]
        colAlpha = openpyxl.utils.get_column_letter(column)
        length = len(mysheet[colAlpha])
        row = sheetColRow[0][2]
        # print(f"len(mysheet[B] = {length = }")
        alphabet_string = string.ascii_uppercase
        alphabet_list = list(alphabet_string)
        titles = finding.exTitles.copy()
        length = len(titles)
        for i in range(1,length):
            assert mysheet.cell(row,i).value == titles[i-1]

    def test_getExcelWords(self,finding,filePathNSheet):
        pathNName,sheetName=filePathNSheet
        sheet = finding.getSheet(pathNName,sheetName)
        # sheetName,col,row
        #(self,val,pathFile,sheetName =None)
        # wordIdTitleEx wordTitleEx wordNumTitleEx
        sheetColRow = finding.findValueInExcel(finding.wordTitleEx,pathNName,sheetName=sheetName)
        # print(f"{sheetColRow = }")
        column =sheetColRow[0][1] 
        columnAlpha = openpyxl.utils.cell.get_column_letter(column)
        # print(f"{columnAlpha = }")

        lengthWords = len(sheet[columnAlpha])-3
        idNwords,words = finding.getExcelWords(pathNName,sheetName)
        # print(f"{idNwords = }")
        assert isinstance(idNwords,list)
        assert isinstance(words,list)
        assert lengthWords==len(words)
        # print(f"{idNwords = }")
        assert lengthWords==len(idNwords) 
        assert len(words)==len(idNwords) 
        for i in range(len(idNwords)):
            
            # print(f"{idNwords[i] = }")
            assert isinstance(idNwords[i],tuple)
            assert len(idNwords[i]) == 2
            # print(f"{idNwords[i] = }")
            assert isinstance(idNwords[i][0],int)
            
            assert isinstance(idNwords[i][1],str) 
            assert isinstance(words[i],str) 
            assert words[i] == idNwords[i][1]
            
        return idNwords
    def test_getWordsSeparated(self,finding,filePathNSheet):
        idNWords = self.test_getExcelWords(finding,filePathNSheet)
        numbers,words = finding.getWordsSeparated(idNWords)
        assert isinstance(words,list)
        assert isinstance(numbers,list)
        for i in range(len(words)):
            # print(f"{words[i] = }")
            assert isinstance(words[i],str)
            # print(f"{numbrs[i] = }")
            assert isinstance(numbers[i],int)
        # print(f"{len(words) = }")
        # print(f"{len(numbers) = }")    
        assert len(words) == len(numbers)

    def test_getMongoWords(self,finding):
        mongoWords = finding.getMongoWords()
        assert isinstance(mongoWords,list)
        return mongoWords

    def test_nonExistanceWordsInMongo(self,finding,filePathNSheet):
        pathNName,sheetName=filePathNSheet
        sheet = finding.getSheet(pathNName,sheetName)
        idNWords = self.test_getExcelWords(finding,filePathNSheet)
        _,words = finding.getWordsSeparated(idNWords)
        mongoWords = self.test_getMongoWords(finding)
        wordList = finding.nonExistanceWordsInMongo(pathNName,sheetName)
        # print(f"{idNWords = }")
        # print(f"\n\n{wordList = }")
        assert isinstance(wordList,list)
        for i in mongoWords:
            assert i not in wordList
    #FIXME test_getFilledWordsFromExcelAllContent    
    @pytest.mark.parametrize("word",[
        ("geben"),
        ("grüß Gott"),
        ("die Mahlzeit"),
        (""),
        ("None")
    ])  
    def test_getFilledWordsFromExcelAllContent(self,word,finding,filePathNSheet):
        if finding.getFilledWordsFromExcelAllContent(word) is not False:
            wordType,article,voiceName,voiceTelLink = finding.getFilledWordsFromExcelAllContent(word)
            wordTypeVerList = ["Verb","Adverb","Adjektiv","Nomen","Possessivpronomen","Interjektion","____"]
            assert wordType in  wordTypeVerList
            assert isinstance(article,str)
            artcleKinds = ["der","die","das","____"]
            assert article in artcleKinds
            assert isinstance(voiceName,str)
            vNList = voiceName.split(".")
            assert vNList[len(vNList)-1] == "mp3"
            assert isinstance(voiceTelLink,str)
            vTLList = voiceTelLink.split("ce/")
            assert vTLList[0]+"ce/" == "https://t.me/guew_resource/"

    @pytest.mark.parametrize("word",[
        ("geben"),
        ("grüß Gott"),
        ("die Mahlzeit"),
        ("tschau"),
        ("servus")
    ])    
    def test_getWordsFromExcelAllContent(self,word,finding,filePathNSheet):
        pathNName,sheetName=filePathNSheet
        sheet = finding.getSheet(pathNName,sheetName)
        if finding.getWordsFromExcelAllContent(word) is not False:
            wordType,article,voiceName,voiceTelLink,meaningId = finding.getWordsFromExcelAllContent(word)
            # print(f"{word = }")
            if wordType != None:
                wordTypeVerList = ["Verb","Adverb","Adjektiv","Nomen","Possessivpronomen","Interjektion","____"]
                assert wordType in  wordTypeVerList
            if article != None:
                assert isinstance(article,str)
                artcleKinds = ["der","die","das","____"]
                assert article in artcleKinds
            if voiceName != None:
                assert isinstance(voiceName,str)
                vNList = voiceName.split(".")
                assert vNList[len(vNList)-1] == "mp3"
            if voiceTelLink !=None:
                assert isinstance(voiceTelLink,str)
                vTLList = voiceTelLink.split("ce/")
                assert vTLList[0]+"ce/" == "https://t.me/guew_resource/"
            if meaningId !=None:
                assert isinstance(meaningId,list)
                if len(meaningId) != 0:
                    for eachId in meaningId:
                        assert isinstance(eachId,str)


    @pytest.mark.parametrize("word",[
        ("sich freuen [vr]"),
        ("die Entschuldigung;-,en"),
        ("heißen [vi]"),
        ("auf wiedersehen")
    ])
    def test_getWordsFromMongoAllContent(self,word,finding,db):

        outputFunc = finding.getWordsFromMongoAllContent(word)
        if len(outputFunc)!=0:
            # wordTypeMongo,articleMongo,voiceNameMongo,voiceTelLinkMongo,meaningIdMongo=outputWordContentMongo[0],outputWordContentMongo[1],outputWordContentMongo[2],outputWordContentMongo[3],outputWordContentMongo[4]
            wordType,article,voiceName,voiceTelLink,meaningId =outputFunc[0],outputFunc[1],outputFunc[2],outputFunc[3],outputFunc[4]
            outputMongo = {"wordType":1,"article":1,"voice":1,"meaningId":1,"_id":0}
            wordTypeTest,articleTest,voiceNameTest,voiceTelLinkTest,meaningIdTest=None,None,None,None,""
            for col in db.wordCollection.find({"word":word},outputMongo):
                for field,val in col.items():
                    if field == "wordType":
                        wordTypeTest = val
                    if field == "article":
                        articleTest = val
                    if field == "voice":
                        # print(f"voice val {val = }")
                        voiceNameTest = val["name"]
                        voiceTelLinkTest = val["telegramLink"]
                        # print(f"{voiceNameTest = }")
                        # print(f"{voiceTelLink = }")
                    if field == "meaningId":
                        meaningIdTest = val   
            # wordType,article,voiceName,voiceTelLink,meaningId
            # print(f"{wordType = }\n{article = }\n{voiceName = }\n{voiceTelLink = }\n{meaningId = }")
            assert wordType == wordTypeTest
            assert article == articleTest
            assert voiceName == voiceNameTest
            assert voiceTelLink == voiceTelLinkTest
            assert meaningId == meaningIdTest


    @pytest.mark.parametrize("val,pathFile",[
        ("sich freuen [vr]",r"D:\project\WoerterbuchProject\mainV2\dataEntry\words.xlsx"),
        ("Good Morning",r"D:\project\WoerterbuchProject\mainV2\dataEntry\words.xlsx"),
        ("hello",r"D:\project\WoerterbuchProject\mainV2\dataEntry\words.xlsx"),
        ("Hallo",r"D:\project\WoerterbuchProject\mainV2\dataEntry\words.xlsx"),
        ("die",r"D:\project\WoerterbuchProject\mainV2\dataEntry\words.xlsx")
    ])
    def test_FindValueInExcel(self,val,pathFile,finding):
        rowColumnList = finding.findValueInExcel(val,pathFile)
        # print(f"{val = }")
        # print(f"{rowColumnList = }")
        assert isinstance(rowColumnList,list)
        for i in rowColumnList:
            assert isinstance(i,tuple)
            assert isinstance(i[0],str)
            assert isinstance(i[1],int)
            assert isinstance(i[2],int)
    #FIXME test_getMeaningWordFromExcelAllDetails
    @pytest.mark.parametrize("word,noneValues",[
        ("der Dank",(None,None)),
        ("sich entschuldigen",("deutschMeaning2",None)),
        ("vorstellen jemanden / sich (jemandem)",(None,456465)),
        ("Ihr",("Ihr Possessivpronomen der höflichen Form der 2. Person Sg und Pl (Sie); Tabellen unter Possessivpronomen und unter mein",45564564)),
        ("der Tag; -(e)s, -e",("deutschMeaning5",5454532))

    ])
    def test_getMeaningWordFromExcelAllDetails(self,word,noneValues,finding):
        #word,deutschMeaning=None,maningId=None
        # print(f"{noneValues = }")
        output = finding.getMeaningWordFromExcelAllDetails(word)
        if len(output) !=0:
            # print(f"{finding.getMeaningWordFromExcelAllDetails(word) = }")
            if noneValues[0] ==None and noneValues[1]==None:
                meaningId,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer = output[0],output[1],output[2],output[3],output[4],output[5],output[6],output[7],output[8],output[9],output[10],output[11],output[12],output[13]
            elif noneValues[0] != None and noneValues[1] == None: 
                output= finding.getMeaningWordFromExcelAllDetails(word,deutschMeaning=noneValues[0])
                meaningId,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer = output[0],output[1],output[2],output[3],output[4],output[5],output[6],output[7],output[8],output[9],output[10],output[11],output[12],output[13]
            elif noneValues[0] == None and noneValues[1] != None:

                output = finding.getMeaningWordFromExcelAllDetails(word,meaningId=noneValues[1])
                meaningId,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer = output[0],output[1],output[2],output[3],output[4],output[5],output[6],output[7],output[8],output[9],output[10],output[11],output[12],output[13]
            elif noneValues[0] != None and noneValues[1] != None: 
                output= finding.getMeaningWordFromExcelAllDetails(word,deutschMeaning=noneValues[0],meaningId=noneValues[1])
                meaningId,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer = output[0],output[1],output[2],output[3],output[4],output[5],output[6],output[7],output[8],output[9],output[10],output[11],output[12],output[13]
            assert isinstance(meaningId,list)
            for eachId in meaningId:
                assert isinstance(eachId,str)
                
            if meaningNumDAF != None:
                assert isinstance(meaningNumDAF,int)
            if deutsch != None:
                assert isinstance(deutsch,str)
            if english !=None:
                assert isinstance(english,str)
            if synonym !=None: 
                assert isinstance(synonym,str)
            if persian != None:
                assert isinstance(persian,str)
            if nameDeAudio != None:
                    assert isinstance(nameDeAudio,str)
                    nDAList = nameDeAudio.split(".")
                    assert nDAList[len(nDAList)-1] == "mp3"
            if addTelDeu != None:
                    assert isinstance(addTelDeu,str)
                    aTDList = addTelDeu.split("ce/")
                    assert aTDList[0]+"ce/" == "https://t.me/guew_resource/"
            if nameSynAudio != None:
                    assert isinstance(nameSynAudio,str)
                    nSAList = nameSynAudio.split(".")
                    assert nSAList[len(nSAList)-1] == "mp3"
            if addTelSyn != None:
                    assert isinstance(addTelSyn,str)
                    aTSList = addTelSyn.split("ce/")
                    assert aTSList[0]+"ce/" == "https://t.me/guew_resource/"
            if nameEnAudio != None:
                    assert isinstance(nameEnAudio,str)
                    nEAList = nameEnAudio.split(".")
                    assert nEAList[len(nEAList)-1] == "mp3"
            if addTelEn != None:
                    assert isinstance(addTelEn,str)
                    aTEList = addTelSyn.split("ce/")
                    assert aTEList[0]+"ce/" == "https://t.me/guew_resource"             
            if namePerAudio != None:
                    assert isinstance(namePerAudio,str)
                    nPAList = namePerAudio.split(".")
                    assert nPAList[len(nPAList)-1] == "mp3"
            if addTelPer != None:
                    assert isinstance(addTelPer,str)
                    aTPList = addTelPer.split("ce/")
                    assert aTPList[0]+"ce/" == "https://t.me/guew_resource"  
    @pytest.mark.parametrize("word,id",[
        ("sich freuen [vr]","5dc5c380ec972f0ee087361d"),
        ("die Entschuldigung;-,en","5dc5c4fdec972f0ee087361e"),
        ("der Name; -ns, -n","5dc5c512ec972f0ee087361f")

    ])
    def test_getMeaningWordFromMongoAllDetails(self,word,id,finding,deutsch=None,meaningId=None):
        # id = Obj(id)
        meaningId,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer = finding.getMeaningWordFromMongoAllDetails(word,ObjectId(id))
        assert meaningId == ObjectId(id)
        # print(type(meaningId))
        assert isinstance(meaningId,bson.objectid.ObjectId)
        assert isinstance(meaningNumDAF,str) or meaningNumDAF  == None
        assert isinstance(deutsch,str)
        assert isinstance(english,str)
        assert isinstance(synonym,str)
        assert isinstance(persian,str)
        if nameDeAudio != None:
            assert isinstance(nameDeAudio,str)
            nDAList = nameDeAudio.split(".")
            assert nDAList[len(nDAList)-1] == "mp3"
        if addTelDeu != None:
                assert isinstance(addTelDeu,str)
                aTDList = addTelDeu.split("ce/")
                assert aTDList[0]+"ce/" == "https://t.me/guew_resource/"
        if nameSynAudio != None:
                assert isinstance(nameSynAudio,str)
                nSAList = nameSynAudio.split(".")
                assert nSAList[len(nSAList)-1] == "mp3"
        if addTelSyn != None:
                assert isinstance(addTelSyn,str)
                aTSList = addTelSyn.split("ce/")
                assert aTSList[0]+"ce/" == "https://t.me/guew_resource/"
        if nameEnAudio != None:
                assert isinstance(nameEnAudio,str)
                nEAList = nameEnAudio.split(".")
                assert nEAList[len(nEAList)-1] == "mp3"
        if addTelEn != None:
                assert isinstance(addTelEn,str)
                aTEList = addTelSyn.split("ce/")
                assert aTEList[0]+"ce/" == "https://t.me/guew_resource/"             
        if namePerAudio != None:
                assert isinstance(namePerAudio,str)
                nPAList = namePerAudio.split(".")
                assert nPAList[len(nPAList)-1] == "mp3"
        if addTelPer != None:
                assert isinstance(addTelPer,str)
                aTPList = addTelPer.split("ce/")
                assert aTPList[0]+"ce/" == "https://t.me/guew_resource/"  



    @pytest.mark.parametrize("word",[
        ("sich freuen [vr]"),
        ("die Entschuldigung;-,en"),
        ("der Name; -ns, -n")
    ])
    def test_getmeaningIdFromExcel(self,word,finding):
        meaningIds = finding.getmeaningIdFromExcel(word)
        assert isinstance(meaningIds,list)
        if len(meaningIds) !=0:
            for x in meaningIds:
                assert isinstance(x,str)
    @pytest.mark.parametrize("word",[
        ("sich freuen [vr]"),
        ("die Entschuldigung;-,en"),
        ("der Name; -ns, -n")
    ])
    def test_getmeaningIdFromMongo(self,word,finding):
        meaningId = finding.getmeaningIdFromMongo(word)
        assert isinstance(meaningId,list)
        if len(meaningId) !=0:
            for x in meaningId:
                # print(type(x))
                assert isinstance(x,bson.objectid.ObjectId)


    @pytest.mark.parametrize("word,noneValues",[
        ("der Dank",(None,None)),
        ("sich entschuldigen",("deutschMeaning2",None)),
        ("vorstellen jemanden / sich (jemandem)",(None,456465)),
        ("Ihr",("Ihr Possessivpronomen der höflichen Form der 2. Person Sg und Pl (Sie); Tabellen unter Possessivpronomen und unter mein",45564564)),
        ("der Tag; -(e)s, -e",("deutschMeaning5",5454532))

    ])
    def test_getFilledMeaningWordAllDetails(self,word,noneValues,finding):
        if finding.getFilledMeaningWordAllDetails(word) is not False:
            if noneValues[0] ==None and noneValues[1]==None:
                meaningId,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer = finding.getFilledMeaningWordAllDetails(word)
            elif noneValues[0] != None and noneValues[1] == None:
                meaningId,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer = finding.getFilledMeaningWordAllDetails(word,deutschMeaning=noneValues[0])
            elif noneValues[0] == None and noneValues[1] != None:
                meaningId,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer = finding.getFilledMeaningWordAllDetails(word,maningId=noneValues[1])
            elif noneValues[0] != None and noneValues[1] != None:
                meaningId,meaningNumDAF,deutsch,english,synonym,persian,nameDeAudio,addTelDeu,nameSynAudio,addTelSyn,nameEnAudio,addTelEn,namePerAudio,addTelPer = finding.getFilledMeaningWordAllDetails(word,deutschMeaning=noneValues[0],maningId=noneValues[1])

            assert isinstance(meaningId,int)
            assert isinstance(meaningNumDAF,int)
            assert isinstance(deutsch,str)
            assert isinstance(english,str)
            assert isinstance(synonym,str)
            assert isinstance(persian,str)
            if nameDeAudio != None:
                    assert isinstance(nameDeAudio,str)
                    nDAList = nameDeAudio.split(".")
                    assert nDAList[len(nDAList)-1] == "mp3"
            if addTelDeu != None:
                    assert isinstance(addTelDeu,str)
                    aTDList = addTelDeu.split("ce/")
                    assert aTDList[0]+"ce/" == "https://t.me/guew_resource/"
            if nameSynAudio != None:
                    assert isinstance(nameSynAudio,str)
                    nSAList = nameSynAudio.split(".")
                    assert nSAList[len(nSAList)-1] == "mp3"
            if addTelSyn != None:
                    assert isinstance(addTelSyn,str)
                    aTSList = addTelSyn.split("ce/")
                    assert aTSList[0]+"ce/" == "https://t.me/guew_resource/"
            if nameEnAudio != None:
                    assert isinstance(nameEnAudio,str)
                    nEAList = nameEnAudio.split(".")
                    assert nEAList[len(nEAList)-1] == "mp3"
            if addTelEn != None:
                    assert isinstance(addTelEn,str)
                    aTEList = addTelSyn.split("ce/")
                    assert aTEList[0]+"ce/" == "https://t.me/guew_resource"             
            if namePerAudio != None:
                    assert isinstance(namePerAudio,str)
                    nPAList = namePerAudio.split(".")
                    assert nPAList[len(nPAList)-1] == "mp3"
            if addTelPer != None:
                    assert isinstance(addTelPer,str)
                    aTPList = addTelPer.split("ce/")
                    assert aTPList[0]+"ce/" == "https://t.me/guew_resource"  
    @pytest.mark.parametrize("word",[
        ("der Dank",),
        ("sich entschuldigen"),
        ("vorstellen jemanden / sich (jemandem)"),
        ("Ihr"),
        ("der Tag; -(e)s, -e")

    ])               
    def test_getBookAllcontentFromExcel(self,finding,word):
            if finding.getBookAllcontentFromExcel(word) is not False:
            # print(f"{finding.getMeaningWordFromExcelAllDetails(word) = }")
                bookName,chapter,content,page,publisher,meaningId = finding.getBookAllcontentFromExcel(word)
                if bookName !=None:
                    assert isinstance(bookName,str)
                if chapter != None:
                    assert isinstance(chapter,str)
                if content != None:
                    assert isinstance(content,str)
                if page !=None:
                    assert isinstance(page,str) or isinstance(page,int)
                if publisher !=None: 
                    assert isinstance(publisher,str)
                if meaningId !=None:
                    assert isinstance(meaningId,str)

    @pytest.mark.parametrize("word",[
        ("der Dank"),
        ("sich entschuldigen"),
        ("vorstellen jemanden / sich (jemandem)"),
        ("Ihr"),
        ("der Tag; -(e)s, -e")

    ])               
    def test_getBookAllcontentFromMongo(self,finding,word):
            output = finding.getBookAllcontentFromMongo(word)
            if len(output) != 0:
            # print(f"{finding.getMeaningWordFromExcelAllDetails(word) = }")

                bookName,chapter,content,page,publisher,meaningId = output[0],output[1],output[2],output[3],output[4],output[5]
                if bookName !=None:
                    assert isinstance(bookName,str)
                if chapter != None:
                    assert isinstance(chapter,str)
                if content != None:
                    assert isinstance(content,str)
                if page !=None:
                    assert isinstance(page,str) or isinstance(page,int)
                if publisher !=None: 
                    assert isinstance(publisher,str)
                if meaningId !=None:
                    assert isinstance(meaningId,str)  


    @pytest.mark.parametrize("word",[
        ("der Dank"),
        ("sich entschuldigen"),
        ("vorstellen jemanden / sich (jemandem)"),
        ("Ihr"),
        ("der Tag; -(e)s, -e")

    ])   
    def test_getFilledBookAllcontent(self,finding,word):
            if finding.getFilledBookAllcontent(word) is not False:
            # print(f"{finding.getMeaningWordFromExcelAllDetails(word) = }")
                bookName,chapter,content,page,publisher = finding.getFilledBookAllcontent(word)
                assert isinstance(bookName,str)
                assert isinstance(chapter,str)
                assert isinstance(content,str)
                assert isinstance(page,str) or isinstance(page,int)
                assert isinstance(publisher,str)
    @pytest.mark.parametrize("word",[
        ("geben"),
        ("selbstverständlich"),
        ("die Karte; -, -n"),
        ("sich melden"),
        ("haben"),
    ])
    def test_getIdFromMongo(self,word,finding,db,filePathNSheet):
        pathNName,sheetName=filePathNSheet
        words = finding.getExcelWords(pathNName,sheetName)
        # # wordIdTitleEx bookIdTitleEx meaningIdTitleEx
        listOrFalse = finding.getIdFromMongo(word)
        # print(f"{listOrFalse = } -- {word = }")
        if isinstance(listOrFalse,list):
            
            assert len(listOrFalse) ==3
            assert None not in listOrFalse
            wordId,bookId,meaningId = None,None,None
            condition = {"word":word}
            output= {"_id":1}
            for col in db.wordCollection.find(condition,output):
                for keys in col.keys():
                    wordId = col[keys]
            assert listOrFalse[0]==wordId
            for col in db.bookCollection.find(condition,output):
                for keys in col.keys():
                    bookId = col[keys]
            assert listOrFalse[1]==bookId
            for col in db.meaningCollection.find(condition,output):
                for keys in col.keys():
                    meaningId = col[keys]
            assert listOrFalse[2]==meaningId




class TestEntering:

    @pytest.mark.parametrize("word,meaningId,wordType,article,voiceName,voiceTelLink",[
        ("test freuen [vr]","5da748126c8002022619ea06","Verb","____","sich freuen_Wort.mp3","https://t.me/guew_resource/277"),
        ("test2","5da748126c8002022619aa06","Nomen","die","some2.mp3","https://t.me/guew_resource/1222"),
        ("test3","5da748126c8002892619ea06","Interjektion","das","some3.mp3","https://t.me/guew_resource/1211"),
        ("test4","5da748126c8982022619ea06","Adverb","der","some4.mp3","https://t.me/guew_resource/1233"),
        ("test5","5da747686c8002022619ea06","____","der","some5.mp3","https://t.me/guew_resource/1244")
    ])
    def test_enterMongoWordAllContents(self,word,meaningId,wordType,article,voiceName,voiceTelLink,entering,db):    
        # word,wordType,article,voiceName,voiceTelLink ="test1","Verb","der","some1.mp3","https://t.me/guew_resource/1234"
        idWordOrFlase = entering.enterMongoWordAllContents(word,meaningId,wordType=wordType,article=article,voiceName=voiceName,voiceTelLink=voiceTelLink)
        if idWordOrFlase is not False:
            assert isinstance(idWordOrFlase,bson.objectid.ObjectId) 
            condition = {"word":word,"meaningId":{"$in":[ObjectId(meaningId)]}}
            output = {"_id":0,"word":1,"wordType":1,"article":1,"voice":1,"meaningId":1}
            oPWord,oPWordType,oPArticle,oPVoiceName,oPVoiceTelLink,meaningIdList=None,None,None,None,None,None
            for col in db.wordCollection.find(condition,output):
                # print(f"{col = }")

                for field,val in col.items():
                    if field == "word":
                        oPWord=val
                    if field == "wordType":
                        oPWordType = val
                    if field == "article":
                        oPArticle = val
                    if field == "voice":
                        oPVoiceName = val["name"]
                        oPVoiceTelLink = val["telegramLink"]
                    if field == "meaningId":
                        meaningIdList=val    

                # print(word,wordType,article,voiceName,voiceTelLink)
                # print(oPWord,oPWordType,oPArticle,oPVoiceName,oPVoiceTelLink)
                for eachId in meaningIdList:
                    assert isinstance(eachId,bson.objectid.ObjectId)
                assert isinstance(oPWord,str)
                assert word == oPWord

                assert isinstance(oPWordType,str)
                assert wordType == oPWordType
                wordTypeVerList = ["Adverb","Verb","Adjektiv","Nomen","Possessivpronomen","Interjektion","____"]
                assert oPWordType in  wordTypeVerList

                assert isinstance(oPArticle,str)
                assert article == oPArticle
                artcleKinds = ["der","die","das","____"]
                assert oPArticle in artcleKinds

                assert isinstance(oPVoiceName,str)
                assert voiceName == oPVoiceName
                vNList = oPVoiceName.split(".")
                assert vNList[len(vNList)-1] == "mp3"

                assert isinstance(oPVoiceTelLink,str)
                assert voiceTelLink == oPVoiceTelLink
                vTLList = oPVoiceTelLink.split("ce/")
                assert vTLList[0]+"ce/" == "https://t.me/guew_resource/"



            # cond = {"word":word,"meaningId":{"$in":ObjectId(meaningId)}}
            verfication = False
            
            cond = {"word":word,"meaningId":{"$in":[ObjectId(meaningId)]},"_id":idWordOrFlase}
            db.wordCollection.delete_one(cond)
            for col in db.wordCollection.find(cond,{"_id":1}):
                for keys in col.keys():
                    verfication = True
            assert verfication is False


    @pytest.mark.parametrize("deutsch,english,synonym,persian,meaningNumDaf,word,voice",[
        ("deutsch1 meaning ","english1 meaning","synonym1 meaning","persian1 meaning",2,"test1",["deuName1.mp3"]),
        ("deutsch2 meaning ","english2 meaning","synonym2 meaning","persian2 meaning",1,"test2",["deuName2.mp3","https://t.me/guew_resource/4545486"]),
        ("deutsch3 meaning ","english3 meaning","synonym3 meaning","persian3 meaning",3,"test3",["deuName3.mp3","https://t.me/guew_resource/456665","engName3.mp3"]),
        ("deutsch4 meaning ","english4 meaning","synonym4 meaning","persian4 meaning",1,"test4",["deuName4.mp3","https://t.me/guew_resource/1465","engName4.mp3","https://t.me/guew_resource/58446"]),
        ("deutsch5 meaning ","english5 meaning","synonym5 meaning","persian5 meaning",1,"test5",["deuName5.mp3","https://t.me/guew_resource/44654","engName5.mp3","https://t.me/guew_resource/5412168","synName5.mp3"]),

        ("deutsch6 meaning ","english6 meaning","synonym6 meaning","persian6 meaning",1,"test6",["deuName6.mp3","https://t.me/guew_resource/416545","engName6.mp3","https://t.me/guew_resource/54513","synName6.mp3","https://t.me/guew_resource/897867687"]),
        ("deutsch7 meaning ","english7 meaning","synonym7 meaning","persian7 meaning",1,"test7",["deuName7.mp3","https://t.me/guew_resource/7545","engName7.mp3","https://t.me/guew_resource/654423","synName7.mp3","https://t.me/guew_resource/334367","perName7.mp3"]),
        ("deutsch8 meaning ","english8 meaning","synonym8 meaning","persian8 meaning",1,"test8",["deuName8.mp3","https://t.me/guew_resource/756433","engName8.mp3","https://t.me/guew_resource/8964380","synName8.mp3","https://t.me/guew_resource/154625","perName8.mp3","https://t.me/guew_resource/46857215"]),
        ("deutsch9 meaning ","english9 meaning","synonym9 meaning","persian9 meaning",1,"test9",[None])
    ])
    def test_enterMeaningWordAllContents(self,word,deutsch,english,synonym,persian,meaningNumDaf,voice,entering,db):
        insVal = {"word":word,"wordType":"wordType","article":"____","voice":{"name":"name","telegramLink":"https://t.me/guew_resource/1234"},"meaningId":"meaningId"}
        # enterMeaningWordAllContents(self,word,deutsch,english=None,synonym=None,persian=None,meaningNumDaf=None,deuName=None,deuTelegramLink=None,engName=None,engTelegramLink=None,synName=None,synTelegramLink=None,perName=None,perTelegramLink=None):
        db.wordCollection.insert_one(insVal)
        wordId = ""
        if len(voice)==1 and voice[0] is not None: 
            idOrFalse = entering.enterMeaningWordAllContents(word,deutsch,english=english,synonym=synonym,persian=persian,meaningNumDaf=meaningNumDaf,deuName=voice[0])
        elif len(voice)== 2:
            idOrFalse = entering.enterMeaningWordAllContents(word,deutsch,english=english,synonym=synonym,persian=persian,meaningNumDaf=meaningNumDaf,deuName=voice[0],deuTelegramLink=voice[1])
        elif len(voice)==3:
            idOrFalse = entering.enterMeaningWordAllContents(word,deutsch,english=english,synonym=synonym,persian=persian,meaningNumDaf=meaningNumDaf,deuName=voice[0],deuTelegramLink=voice[1],engName=voice[2])    
        elif len(voice)==4:
            idOrFalse = entering.enterMeaningWordAllContents(word,deutsch,english=english,synonym=synonym,persian=persian,meaningNumDaf=meaningNumDaf,deuName=voice[0],deuTelegramLink=voice[1],engName=voice[2],engTelegramLink=voice[3]) 
        elif len(voice)==5:
            idOrFalse = entering.enterMeaningWordAllContents(word,deutsch,english=english,synonym=synonym,persian=persian,meaningNumDaf=meaningNumDaf,deuName=voice[0],deuTelegramLink=voice[1],engName=voice[2],engTelegramLink=voice[3],synName=voice[4])        
        elif len(voice)==6:
            idOrFalse = entering.enterMeaningWordAllContents(word,deutsch,english=english,synonym=synonym,persian=persian,meaningNumDaf=meaningNumDaf,deuName=voice[0],deuTelegramLink=voice[1],engName=voice[2],engTelegramLink=voice[3],synName=voice[4],synTelegramLink=voice[5])
        elif len(voice)==7:
            idOrFalse = entering.enterMeaningWordAllContents(word,deutsch,english=english,synonym=synonym,persian=persian,meaningNumDaf=meaningNumDaf,deuName=voice[0],deuTelegramLink=voice[1],engName=voice[2],engTelegramLink=voice[3],synName=voice[4],synTelegramLink=voice[5],perName=voice[6])
        elif len(voice)==8:
            idOrFalse = entering.enterMeaningWordAllContents(word,deutsch,english=english,synonym=synonym,persian=persian,meaningNumDaf=meaningNumDaf,deuName=voice[0],deuTelegramLink=voice[1],engName=voice[2],engTelegramLink=voice[3],synName=voice[4],synTelegramLink=voice[5],perName=voice[6],perTelegramLink=voice[7]) 
        elif len(voice)==1 and voice[0] is None:
            idOrFalse = entering.enterMeaningWordAllContents(word,deutsch,english=english,synonym=synonym,persian=persian,meaningNumDaf=meaningNumDaf)


        if idOrFalse is not False:

            #FIXed idOrFalse
            assert isinstance(idOrFalse,bson.objectid.ObjectId)
            condition = {"deutsch":deutsch,"english":english,"synonym":synonym,"persian":persian,"meaningNumDaf":meaningNumDaf,"word":word}
            # output = {"_id":0,"deutsch":1,"english":1,"synonym":1,"persian":1,"meaningNumDaf":1,"word":1,}
            output = {"_id":0,"deutsch":1,"english":1,"synonym":1,"persian":1,"meaningNumDaf":1,"word":1,"synVoice":1,"engVoice":1,"perVoice":1,"deuVoice":1}
            oPDeutsch,oPEnglish,oPSynonym,oPPersian,oPMeaningNumDaf,oPWord="","","","","",""

            oPdeuName,oPdeuTelegramLink,oPengName,oPengTelegramLink,oPsynName,oPsynTelegramLink,oPperName,oPperTelegramLink="","","","","","","",""
            for col in db.meaningCollection.find(condition,output):
                # print(f"{col = }")

                for field,val in col.items():
                    if field == "deutsch":
                        oPDeutsch=val
                    if field == "english":
                        oPEnglish = val
                    if field == "synonym":
                        oPSynonym = val
                    if field == "persian":
                        oPPersian = val
                    if field == "meaningNumDaf":
                        oPMeaningNumDaf = val
                    if field == "word":
                        oPWord = val
                        # synVoice,engVoice,perVoice,deuVoice 
                    if len(voice)==1 and field == "deuVoice" and voice[0] is not None: 
                            oPdeuName = val["deuName"]
                            assert oPdeuName == voice[0]
                            dENaList = oPdeuName.split(".")
                            assert dENaList[len(dENaList)-1] == "mp3"


                    elif len(voice)== 2:
                        if field == "deuVoice":
                            oPdeuName = val["deuName"]
                            assert oPdeuName == voice[0]
                            dENaList = oPdeuName.split(".")
                            assert dENaList[len(dENaList)-1] == "mp3"

                            oPdeuTelegramLink = val["deuTelegramLink"]
                            assert oPdeuTelegramLink == voice[1]
                            dETeLiList = oPdeuTelegramLink.split("ce/")
                            assert dETeLiList[0]+"ce/" == "https://t.me/guew_resource/"

                    elif len(voice)==3:
                        if field == "deuVoice":
                            oPdeuName = val["deuName"]
                            assert oPdeuName == voice[0]
                            dENaList = oPdeuName.split(".")
                            assert dENaList[len(dENaList)-1] == "mp3"

                            oPdeuTelegramLink = val["deuTelegramLink"]
                            assert oPdeuTelegramLink == voice[1]
                            dETeLiList = oPdeuTelegramLink.split("ce/")
                            assert dETeLiList[0]+"ce/" == "https://t.me/guew_resource/"
                        elif field == "engVoice":
                            oPengName = val["engName"]
                            assert oPengName == voice[2]
                            eNNaList = oPdeuName.split(".")
                            assert eNNaList[len(eNNaList)-1] == "mp3"
                    elif len(voice)==4:
                        if field == "deuVoice":
                            oPdeuName = val["deuName"]
                            assert oPdeuName == voice[0]
                            dENaList = oPdeuName.split(".")
                            assert dENaList[len(dENaList)-1] == "mp3"

                            oPdeuTelegramLink = val["deuTelegramLink"]
                            assert oPdeuTelegramLink == voice[1]
                            dETeLiList = oPdeuTelegramLink.split("ce/")
                            assert dETeLiList[0]+"ce/" == "https://t.me/guew_resource/"
                        if field == "engVoice":
                            oPengName = val["engName"]
                            assert oPengName == voice[2]
                            eNNaList = oPdeuName.split(".")
                            assert eNNaList[len(eNNaList)-1] == "mp3"

                            oPengTelegramLink = val["engTelegramLink"]
                            assert oPengTelegramLink == voice[3]
                            eNNaList = oPengTelegramLink.split("ce/")
                            assert eNNaList[0]+"ce/" == "https://t.me/guew_resource/"
                    elif len(voice)==5:
                        if field == "deuVoice":
                            oPdeuName = val["deuName"]
                            assert oPdeuName == voice[0]
                            dENaList = oPdeuName.split(".")
                            assert dENaList[len(dENaList)-1] == "mp3"

                            oPdeuTelegramLink = val["deuTelegramLink"]
                            assert oPdeuTelegramLink == voice[1]
                            dETeLiList = oPdeuTelegramLink.split("ce/")
                            assert dETeLiList[0]+"ce/" == "https://t.me/guew_resource/"
                        if field == "engVoice":
                            oPengName = val["engName"]
                            assert oPengName == voice[2]
                            eNNaList = oPdeuName.split(".")
                            assert eNNaList[len(eNNaList)-1] == "mp3"

                            oPengTelegramLink = val["engTelegramLink"]
                            assert oPengTelegramLink == voice[3]
                            eNNaList = oPengTelegramLink.split("ce/")
                            assert eNNaList[0]+"ce/" == "https://t.me/guew_resource/"
                        if field == "synVoice":
                            oPsynName = val["synName"]
                            assert oPsynName == voice[4]
                            sYNaList = oPsynName.split(".")
                            assert sYNaList[len(sYNaList)-1] == "mp3"

                    elif len(voice)==6:
                        if field == "deuVoice":
                            oPdeuName = val["deuName"]
                            assert oPdeuName == voice[0]
                            dENaList = oPdeuName.split(".")
                            assert dENaList[len(dENaList)-1] == "mp3"

                            oPdeuTelegramLink = val["deuTelegramLink"]
                            assert oPdeuTelegramLink == voice[1]
                            dETeLiList = oPdeuTelegramLink.split("ce/")
                            assert dETeLiList[0]+"ce/" == "https://t.me/guew_resource/"
                        if field == "engVoice":
                            oPengName = val["engName"]
                            assert oPengName == voice[2]
                            eNNaList = oPdeuName.split(".")
                            assert eNNaList[len(eNNaList)-1] == "mp3"

                            oPengTelegramLink = val["engTelegramLink"]
                            assert oPengTelegramLink == voice[3]
                            eNNaList = oPengTelegramLink.split("ce/")
                            assert eNNaList[0]+"ce/" == "https://t.me/guew_resource/"
                        if field == "synVoice":
                            oPsynName = val["synName"]
                            assert oPsynName == voice[4]
                            sYNaList = oPsynName.split(".")
                            assert sYNaList[len(sYNaList)-1] == "mp3"

                            oPsynTelegramLink = val["synTelegramLink"]
                            assert oPsynTelegramLink == voice[5]
                            sYTelLiList = oPsynTelegramLink.split("ce/")
                            assert sYTelLiList[0]+"ce/" == "https://t.me/guew_resource/"

                    elif len(voice)==7:
                        if field == "deuVoice":
                            oPdeuName = val["deuName"]
                            assert oPdeuName == voice[0]
                            dENaList = oPdeuName.split(".")
                            assert dENaList[len(dENaList)-1] == "mp3"

                            oPdeuTelegramLink = val["deuTelegramLink"]
                            assert oPdeuTelegramLink == voice[1]
                            dETeLiList = oPdeuTelegramLink.split("ce/")
                            assert dETeLiList[0]+"ce/" == "https://t.me/guew_resource/"
                        if field == "engVoice":
                            oPengName = val["engName"]
                            assert oPengName == voice[2]
                            eNNaList = oPdeuName.split(".")
                            assert eNNaList[len(eNNaList)-1] == "mp3"

                            oPengTelegramLink = val["engTelegramLink"]
                            assert oPengTelegramLink == voice[3]
                            eNNaList = oPengTelegramLink.split("ce/")
                            assert eNNaList[0]+"ce/" == "https://t.me/guew_resource/"
                        if field == "synVoice":
                            oPsynName = val["synName"]
                            assert oPsynName == voice[4]
                            sYNaList = oPsynName.split(".")
                            assert sYNaList[len(sYNaList)-1] == "mp3"

                            oPsynTelegramLink = val["synTelegramLink"]
                            assert oPsynTelegramLink == voice[5]
                            sYTelLiList = oPsynTelegramLink.split("ce/")
                            assert sYTelLiList[0]+"ce/" == "https://t.me/guew_resource/"
                        if field == "perVoice":
                            oPPerName = val["perName"]
                            assert oPPerName == voice[6]
                            pENaList = oPPerName.split(".")
                            assert pENaList[len(pENaList)-1] == "mp3"
                        # oPdeuName,oPdeuTelegramLink,oPengName,oPengTelegramLink,oPsynName,oPsynTelegramLink,oPperName,oPperTelegramLink
                    elif len(voice)==8: 
                        if field == "deuVoice":
                            oPdeuName = val["deuName"]
                            assert oPdeuName == voice[0]
                            dENaList = oPdeuName.split(".")
                            assert dENaList[len(dENaList)-1] == "mp3"

                            oPdeuTelegramLink = val["deuTelegramLink"]
                            assert oPdeuTelegramLink == voice[1]
                            dETeLiList = oPdeuTelegramLink.split("ce/")
                            assert dETeLiList[0]+"ce/" == "https://t.me/guew_resource/"
                        if field == "engVoice":
                            oPengName = val["engName"]
                            assert oPengName == voice[2]
                            eNNaList = oPdeuName.split(".")
                            assert eNNaList[len(eNNaList)-1] == "mp3"

                            oPengTelegramLink = val["engTelegramLink"]
                            assert oPengTelegramLink == voice[3]
                            eNNaList = oPengTelegramLink.split("ce/")
                            assert eNNaList[0]+"ce/" == "https://t.me/guew_resource/"
                        if field == "synVoice":
                            oPsynName = val["synName"]
                            assert oPsynName == voice[4]
                            sYNaList = oPsynName.split(".")
                            assert sYNaList[len(sYNaList)-1] == "mp3"

                            oPsynTelegramLink = val["synTelegramLink"]
                            assert oPsynTelegramLink == voice[5]
                            sYTelLiList = oPsynTelegramLink.split("ce/")
                            assert sYTelLiList[0]+"ce/" == "https://t.me/guew_resource/"
                        if field == "perVoice":
                            oPPerName = val["perName"]
                            assert oPPerName == voice[6]
                            pENaList = oPPerName.split(".")
                            assert pENaList[len(pENaList)-1] == "mp3"


                            oPPerTelegramLink = val["perTelegramLink"]
                            assert oPPerTelegramLink == voice[7]
                            pETelLiList = oPPerTelegramLink.split("ce/")
                            assert pETelLiList[0]+"ce/" == "https://t.me/guew_resource/"
                        
                assert isinstance(oPDeutsch,str)
                assert deutsch == oPDeutsch
                assert isinstance(oPEnglish,str)
                assert english == oPEnglish
                assert isinstance(oPSynonym,str)
                assert synonym == oPSynonym
                assert isinstance(oPPersian,str)
                assert persian == oPPersian
                assert meaningNumDaf == oPMeaningNumDaf
                # assert oPMeaningNumDaf
                assert isinstance(oPWord,str)
                assert word == oPWord
            # print(f"end of method {idOrFalse = }")
        ## test meaning id in word collection
            meaningIdlistWordCollect = []
            # wordId = ""
            for col in db.wordCollection.find({"word":word},{"meaningId":1,"_id":1}):
                # tempmeaningIdlistWordCollect = ""
                for field,val in col.items():
                    tempmeaningIdlistWordCollect = None
                    if field == "meaningId":
                        tempmeaningIdlistWordCollect = val
                    if field == "_id":
                        wordId = val
                meaningIdlistWordCollect = meaningIdlistWordCollect + tempmeaningIdlistWordCollect
            assert idOrFalse in meaningIdlistWordCollect

                

        # print(f"{idOrFalse = }")
        db.meaningCollection.delete_one({"_id":idOrFalse})
        testList = []
        for col in db.meaningCollection.find({"_id":idOrFalse},{"_Id":1}):
            for keys in col.keys():
                testList.append(col[keys])
        assert len(testList)==0
        db.wordCollection.delete_one({"_id":wordId})
        test2List = []
        for col in db.wordCollection.find({"_id":idOrFalse},{"_Id":1}):
            for keys in col.keys():
                test2List.append(col[keys])
        assert len(test2List)==0


    


    @pytest.mark.parametrize("word,bookName,chapter,content,page,publisher,meaningId",[
        ("word","bookName","chapter","content","page","publisher",ObjectId("5dc5c380ec972f0ee087361d")),
        ("word","bookName","chapter","content","page","publisher",ObjectId("5dc5c51cec972f0ee0873622")),
        ("word","bookName","chapter","content","page","publisher",ObjectId("5dc5c525ec972f0ee0873625")),
        ("word","bookName","chapter","content","page","publisher",ObjectId("5dc5c536ec972f0ee087362b")),
        ("word","bookName","chapter","content","page","publisher",ObjectId("5dc5cb79ec972f0ee0873637"))
        
    ])
    def test_enterMongoBookAllContents(self,word,bookName,chapter,content,page,publisher,meaningId,entering,db):
        # bookName,chapter,content,page,publisher
        idOrFalse = entering.enterMongoBookAllContents(word,meaningId,bookName=bookName,chapter=chapter,content=content,page=page,publisher=publisher)
        # print(f"{idOrFalse = }")
        if idOrFalse is not False:
            outputMongo = {"_id":0,"bookName":1,"chapter":1,"content":1,"page":1,"publisher":1,"meaningId":1}
            condition = {"word":word}
            for col in db.bookCollection.find(condition,outputMongo):
                for field,value in col.items():
                    if field == "bookName":
                        assert value == bookName
                        assert isinstance(value,str)
                    if field == "chapter":
                        assert value == chapter
                        assert isinstance(value,str)
                    if field == "content":
                        assert value == content
                        assert isinstance(value,str)
                    if field == "page":
                        assert value == page
                        assert isinstance(value,str) or isinstance(value,int)
                    if field == "publisher":
                        assert value == publisher
                        assert isinstance(value,str)
                    if field == "meaningId":
                        assert value == meaningId
                        assert isinstance(value,bson.objectid.ObjectId)
            db.bookCollection.delete_one({"_id":idOrFalse})

class TestUpdating:
    @pytest.mark.parametrize("word",[
        ("geben"),
        ("selbstverständlich"),
        ("die Karte; -, -n"),
        ("sich melden"),
        ("haben")
    ])
    def test_updateOneIdsWordBookMeaningExcel(self,word,finding,updating,filePathNSheet):
        pathNName,sheetName=filePathNSheet
        
        # ids variable wordIdTitleEx bookIdTitleEx meaningIdTitleEx
        idsList = finding.getIdFromMongo(word)
        # if idsList is not False:
        mySheet = finding.getSheet(pathNName,sheetName)
        output = updating.updateOneIdsWordBookMeaningExcel(mySheet,word,idsList,pathNName,sheetName)
        if output is not False:
           # findValueInExcel(self,val,pathFile,sheetName =None):
           wordSheetColRow = finding.findValueInExcel(word,pathNName,sheetName= sheetName)
        #    print(f"{wordSheetColRow = }") 
           row = wordSheetColRow[0][2]
           wordIdSheetColRow = finding.findValueInExcel(finding.wordIdTitleEx,pathNName,sheetName= sheetName)
           columnWordId = wordIdSheetColRow[0][1]
           bookIdSheetColRow = finding.findValueInExcel(finding.bookIdTitleEx,pathNName,sheetName= sheetName)
           columnBookId = bookIdSheetColRow[0][1]
           meaningIdSheetColRow = finding.findValueInExcel(finding.meaningIdTitleEx,pathNName,sheetName= sheetName)
           columnmeaningId = meaningIdSheetColRow[0][1]
           #    mySheet = finding.getSheet(pathNName,sheetName)
           wordIdVal = mySheet.cell(row,columnWordId).value
           bookIdVal = mySheet.cell(row,columnBookId).value
           meaningIdVal = mySheet.cell(row,columnmeaningId).value
           assert isinstance(wordIdVal,str)
           assert isinstance(bookIdVal,str)
           assert isinstance(meaningIdVal,str)
        #    print(f"{word = }")
           assert idsList[0] == ObjectId(wordIdVal)
           assert idsList[1] == ObjectId(bookIdVal)
           assert idsList[2] == ObjectId(meaningIdVal)

    def test_updateAllIdsWordBookMeaningExcel(self,finding,updating,filePathNSheet):
        pathNName,sheetName=filePathNSheet
        # words = finding.getExcelWords(pathNName,sheetName)
        wordsListTuple = updating.updateAllIdsWordBookMeaningExcel(pathNName,sheetName)
        # if wordsOrFalse is not False:
        for i in wordsListTuple:
            if i[1]==True:
                idsList = finding.getIdFromMongo(i)
                if idsList:
                    wordSheetColumnRow = finding.findValueInExcel(i[0],pathNName,sheetName= sheetName)
                    # print(f"{i[0] = }\n{wordSheetColumnRow = }")
                    row = wordSheetColumnRow[0][2]
                    wordIdSheetColRow = finding.findValueInExcel(finding.wordIdTitleEx,pathNName,sheetName= sheetName)
                    columnWordId = wordIdSheetColRow[0][1]
                    bookIdSheetColRow = finding.findValueInExcel(finding.bookIdTitleEx,pathNName,sheetName= sheetName)
                    columnBookId = bookIdSheetColRow[0][1]
                    meaningIdSheetColRow = finding.findValueInExcel(finding.meaningIdTitleEx,pathNName,sheetName= sheetName)
                    columnmeaningId = meaningIdSheetColRow[0][1]
                    mySheet = finding.getSheet(pathNName,sheetName)
                    wordIdVal = mySheet.cell(row,columnWordId).value
                    bookIdVal = mySheet.cell(row,columnBookId).value
                    meaningIdVal = mySheet.cell(row,columnmeaningId).value
                    print(f"{wordIdVal = }\n{bookIdVal = }\n{meaningIdVal = }\n{idsList = }")
                    assert wordIdVal == idsList[0]
                    assert bookIdVal == idsList[1]
                    assert meaningIdVal == idsList[2]
            else:
                print(f"ids which not has id in excell --> {i}")
    @pytest.mark.parametrize("title,idValue,word",[
        ("word id",ObjectId("5da748126c8002022619ea06"),"sich freuen [vr]"),
        ("book id",ObjectId("5da7467a6c8002022619ea05"),"sich freuen [vr]"),
        ("meaning id",ObjectId("5dc5c380ec972f0ee087361d"),"sich freuen [vr]"),
        ("word id",ObjectId("5da7534f3f99ca0226891b39"),"die Entschuldigung;-,en"),
        ("meaning id",ObjectId("5dc5c4fdec972f0ee087361e"),"die Entschuldigung;-,en")
    ])
    def test_updateExcelId(self,updating,title,idValue,word,finding,filePathNSheet):
        output = False
        pathNName,sheetName=filePathNSheet
        mySheet = finding.getSheet(pathNName,sheetName)
        
        if title =="word id":
            output = updating.updateExcelId(mySheet,title,idValue,word,pathNName,sheetName)
        elif title == "book id":
            output = updating.updateExcelId(mySheet,title,idValue,word,pathNName,sheetName)
        elif title == "meaning id":
            output = updating.updateExcelId(mySheet,title,idValue,word,pathNName,sheetName)

        if output is True:

            titleSheetColRow = finding.findValueInExcel(title,pathNName,sheetName=sheetName)
            column = titleSheetColRow[0][1]
            
            wordSheetColRow = finding.findValueInExcel(word,pathNName,sheetName=sheetName)
            wordRow = wordSheetColRow[0][2]

            idSheetColRow = finding.findValueInExcel(idValue,pathNName,sheetName=sheetName)
            wordIdRow = wordSheetColRow[0][2]
            assert wordRow == wordIdRow
            mySheet = finding.getSheet(pathNName,sheetName)
            idImported = mySheet.cell(wordRow,column).value
            # print(f"{column = }\n{wordRow = }")
            # print(f"{idImported = }")
            # print(f"{idValue =}")
            assert ObjectId(idImported) == idValue
            assert isinstance(idImported,str)

    @pytest.mark.parametrize("wordId,word,wordType,article,meaningId,voiceName,voiceTelLink",[
        ("5f1d6ad1f0abf41950f5a0a5","sich freuen [vr]","Verb","____","5dc5c380ec972f0ee087361d",None,"https://t.me/guew_resource/1234"),
        ("5da7534f3f99ca0226891b39","die Entschuldigung;-,en","Nomen","die","5dc5c4fdec972f0ee087361e","some2.mp3","https://t.me/guew_resource/1222"),
        ("5da755173f99ca0226891b3c","der Herr;-(e)n, -en","Nomen","der","5dc5c752ec972f0ee0873633","some3.mp3","https://t.me/guew_resource/1211"),
        ("5db2cf88a3dcd8031a079a00","heißen [vi]","Verb","____","5dc5cae6ec972f0ee0873635","some4.mp3",None),
        ("5db2cf92a3dcd8031a079a01","sein [vi]","Verb","____","5dc5cae8ec972f0ee0873636",None,None)
    ])
    # (self,word,wordType,article,idsEx,voiceName=None,voiceTelLink=None):
    def test_updateWordDetailsMongo(self,wordId,word,wordType,article,meaningId,voiceName,voiceTelLink,updating,db):
        # updateWordDetailsMongo(self,wordId,word=None,meaningId=None,wordType=None,article=None,voiceName=None,voiceTelLink=None):
        output = updating.updateWordDetailsMongo(wordId,word=word,meaningId=meaningId,wordType=wordType,article=article,voiceName=voiceName,voiceTelLink=voiceTelLink)
        assert isinstance(output,list)
        assert len(output)==6

            
        # print(f"{output = }")
        # if output is True:

        condition = {"_id":ObjectId(wordId)}
        outputMongo = {"_id":0,"word":1,"wordType":1,"article":1,"voice":1,"meaningId":1}
        oPWord,oPWordType,oPArticle,oPVoiceName,oPVoiceTelLink,oPmeaningId=None,None,None,None,None,None
        # print(f"in -->{output = }")
        for col in db.wordCollection.find(condition,outputMongo):
            # print(f"{col = }")

            for field,val in col.items():
                if field == "word":
                    oPWord=val
                if field == "wordType":
                    oPWordType = val
                if field == "article":
                    oPArticle = val
                if field == "voice":
                    oPVoiceName = val["name"]
                    oPVoiceTelLink = val["telegramLink"]
                if field == "meaningId":
                    oPmeaningId = val
        # print(f"{oPmeaningId = }")
        # print(f"{word = }")

        if word == None:
            assert output[0] == None
        else:
            assert isinstance(oPWord,str)
            assert word == oPWord
        if meaningId == None:
            assert output[1] == None
        else:
            assert isinstance(oPmeaningId,list)
            for eachIdMeaningOp in oPmeaningId:
                assert isinstance(eachIdMeaningOp,bson.objectid.ObjectId)
            if meaningId != None:
                assert ObjectId(meaningId) in oPmeaningId

        if wordType == None:
            assert output[2] == None
        else:
            assert wordType == oPWordType
            assert isinstance(oPWordType,str)
            wordTypeVerList = ["Adverb","Verb","Adjektiv","Nomen","Possessivpronomen","Interjektion","____"]
            assert oPWordType in  wordTypeVerList
        if article == None:
            assert output[3] == None
        else:
            assert article == oPArticle
            assert isinstance(oPArticle,str)
            artcleKinds = ["der","die","das","____"]
            assert oPArticle in artcleKinds
        if voiceName == None:
            assert output[4] == None
        else:
            assert voiceName == oPVoiceName
            assert isinstance(oPVoiceName,str)
            vNList = oPVoiceName.split(".")
            assert vNList[len(vNList)-1] == "mp3"
        if voiceTelLink == None:
            assert output[5] == None
        else:
            assert voiceTelLink == oPVoiceTelLink
            assert isinstance(oPVoiceTelLink,str)
            vTLList = oPVoiceTelLink.split("ce/")
            assert vTLList[0]+"ce/" == "https://t.me/guew_resource/"






       
    @pytest.mark.parametrize("wordDetails,deutsch,english,synonym,persian,meaningNumDaf,voice",[
        
        (["word1","wordType1","article1"],"deutsch meaning1","english meaning1","synonym meaning1","persian meaning1",1,["deuName1.mp3","https://t.me/guew_resource/1545644","enName1.mp3","https://t.me/guew_resource/1544344","synName1.mp3","https://t.me/guew_resource/15344","perName1.mp3","https://t.me/guew_resource/13244"]),

        (["word2","wordType2","article2"],"deutsch meaning2","english meaning2","synonym meaning2","persian meaning3",2,["deuName2.mp3",None,None,None,None,None,"perName2.mp3","https://t.me/guew_resource/13244"]),

        (["word3","wordType3","article3"],"deutsch meaning3","english meaning3","synonym meaning3","persian meaning3",3,["deuName3.mp3",None,None,"synName3.mp3","https://t.me/guew_resource/15344",None,None,None])

    ])
    def test_updateMeaningDetailsMongo(self,deutsch,english,synonym,persian,meaningNumDaf,voice,wordDetails,updating,db,entering):
        doc ={"wordDetails":{"word":"test","wordType":"","article":""},"deutsch":1,"english":1,"synonym":1,"persian":1,"meaningNumDaf":1,"deuVoice":{"deuName":"","deuTelegramLink":""},"engVoice":{"engName":"","engTelegramLink":""},"synVoice":{"synName":"","synTelegramLink":""},"perVoice":{"perName":"","perTelegramLink":""}}
        db.meaningCollection.insert_one(doc)
        cond = {"wordDetails":{"word":"test","wordType":"","article":""}}
        meaningId = None
        for col in db.meaningCollection.find(cond,{"_id":1}):
            for keys in col.keys():
                meaningId = col[keys]


        # print(f"{meaningId = }")
        output = updating.updateMeaningDetailsMongo(meaningId,word=wordDetails[0],wordType=wordDetails[1],article=wordDetails[2],deutsch=deutsch,english=english,synonym=synonym,persian=persian,meaningNumDaf=meaningNumDaf,deuName=voice[0],deuTelegramLink=voice[1],engName=voice[2],engTelegramLink=voice[3],synName=voice[4],synTelegramLink=voice[5],perName=voice[6],perTelegramLink=voice[7]) 
        print(f"{output = }")
        assert isinstance(output,list)
        assert len(output)==16
        inputCheck = [wordDetails[0],wordDetails[1],wordDetails[2],deutsch,english,synonym,persian,meaningNumDaf,voice[0],voice[1],voice[2],voice[3],voice[4],voice[5],voice[6],voice[7]]
        testCheck= [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None]
        fields = ["word","wordType","article","deutsch","english","synonym","persian","meaningNumDaf","deuName","deuTelegramLink","engName","engTelegramLink","synName","synTelegramLink","perName","perTelegramLink"]
        condition = {"_id":meaningId}
        ouputMongo = {"wordDetails":1,"deutsch":1,"english":1,"synonym":1,"persian":1,"meaningNumDaf":1,"deuVoice":1,"engVoice":1,"synVoice":1,"perVoice":1,"_id":0}
        verificationId = False
        for col in db.meaningCollection.find(condition,ouputMongo):
           for field,val in col.items():
                verificationId = True
                if field == "wordDetails":
                   testCheck[0]=val[fields[0]]
                   testCheck[1]=val[fields[1]]
                   testCheck[2]=val[fields[2]]
                if field == fields[3]:
                   testCheck[3]=val 
                if field == fields[4]:
                   testCheck[4]=val
                if field == fields[5]:
                   testCheck[5]=val
                if field == fields[6]:
                   testCheck[6]=val
                if field == fields[7]:
                   testCheck[7]=val
                if field == "deuVoice":
                   testCheck[8]=val[fields[8]]
                   testCheck[9]=val[fields[9]] 
                if field == "engVoice":
                   testCheck[10]=val[fields[10]] 
                   testCheck[11]=val[fields[11]]
                if field == "synVoice":
                   testCheck[12]=val[fields[12]] 
                   testCheck[13]=val[fields[13]] 
                if field == "perVoice":
                   testCheck[14]=val[fields[14]] 
                   testCheck[15]=val[fields[15]]
        if verificationId is True:
            for x in range(len(inputCheck)):
                assert isinstance(output[x],bool)
                if inputCheck[x] != None:
                    if output[x] == True:                       
                        assert inputCheck[x] == testCheck[x]
                    else:
                        assert output[x] == False
        db.meaningCollection.delete_one(condition)
                       


    @pytest.mark.parametrize("word,bookName,chapter,content,page,publisher,meaningId",[
        # ("5da7467a6c8002022619ea05","sich freuen [vr]","Großes Übungsbuch Wortschatz","1 sich vorstellen 1 sich vorstellen","A-Kontakte, Informationen zur Person A-Kontakte, Informationen zur Person","8&9","Hueber",ObjectId("5da7467a6c8002022619ea05")),
        # ("5da754ad3f99ca0226891b3b","die Entschuldigung;-,en","Großes Übungsbuch Wortschatz","1 sich vorstellen","A-Kontakte, o hallo","8&9","Hueber Hueber",ObjectId("5dc5c4fdec972f0ee087361e")),
        # ("5da754ad3f99ca0226891b3c","word1","bookName","chapter","content","page","publisher",ObjectId("5dc5c525ec972f0ee0873625")),
        # # this two down make changes back to normal
        # ("5da7467a6c8002022619ea05","sich freuen [vr]","Großes Übungsbuch Wortschatz","1 sich vorstellen 1 sich vorstellen","A-Kontakte, Informationen zur Person","8&9","Hueber",ObjectId("5da7467a6c8002022619ea05")),
        # ("5da754ad3f99ca0226891b3b","die Entschuldigung;-,en","Großes Übungsbuch Wortschatz","1 sich vorstellen","A-Kontakte, Informationen zur Person","8&9","Hueber",ObjectId("5dc5c4fdec972f0ee087361e")),
        ("test1","book1","chapter1","content1","page1","publisher1",ObjectId("5e5c0c5164e446abc050db16")),
        ("test2","book2","chapter2","content2","page2","publisher2",ObjectId("5e5c0c5f64e446abc050db17")),
        ("test3","book3","chapter3","content3","page3","publisher3",ObjectId("5e5c093264e446abc050db00")),
        ("test4","book4","chapter4","content4","page4","publisher4",ObjectId("5e5c08ef64e446abc050daff")),
        ("test5","book5","chapter5","content5","page5","publisher5",ObjectId("5e5c08a064e446abc050dafe"))
        
    ])
    def test_updateBookDetailsMongo(self,word,bookName,chapter,content,page,publisher,meaningId,updating,db):
        insertTest = {"word":word,"bookName":bookName,"chapter":chapter,"content":content,"page":page,"publisher":publisher,"meaningId":meaningId}


        db.bookCollection.insert_one(insertTest)
        bookId = None
        for col in db.bookCollection.find(insertTest,{"_id":1}):
            for keys in col.keys():
                bookId = col[keys]


        output = updating.updateBookDetailsMongo(bookId,word=word,bookName="Hallo book Name",chapter=chapter,content=content,page=page,publisher=publisher,meaningId=meaningId)
        print(f"{output = }")
        assert isinstance(output,list)
        assert len(output)==7
        # opWord,opBookName,opChapter,opContent,opPage,opPublisher,opMeaningId=None,None,None,None,None,None,None
        inputCheck =[word,"Hallo book Name",chapter,content,page,publisher,meaningId] 
        getCheck=[None,None,None,None,None,None,None]
        changedBookId = None
        if isinstance(bookId,str):
            changedBookId = ObjectId(bookId)
        else:
            changedBookId = bookId
        condition = {"_id":changedBookId}
        outputMongo = {"_id":0,"word":1,"bookName":1,"chapter":1,"content":1,"page":1,"publisher":1,"meaningId":1}
        for col in db.bookCollection.find(condition,outputMongo):
            for field,val in col.items():
                if field == "word":
                    # opWord = val
                    getCheck[0]= val
                if field == "bookName":
                    # opBookName = val 
                    getCheck[1]= val
                if field == "chapter":
                    # getCheck = val
                    getCheck[2]= val
                if field == "content":
                    # opContent = val
                    getCheck[3]= val
                if field == "page":
                    # opPage = val
                    getCheck[4]= val 
                if field == "publisher":
                    # opPublisher = val
                    getCheck[5]= val
                if field == "meaningId":
                    # opMeaningId = val
                    getCheck[6]= val
        for x in range(len(getCheck)):
            # print(f"{x = }\n{}")
            assert isinstance(output[x],bool)
            if getCheck[x] !=None:
                # print(f"{x = }\n{getCheck[x] = }")
                if output[x] == True:
                    # assert output[x] == True
                    assert getCheck[x] == inputCheck[x]
                if x == 6 :
                    assert isinstance(getCheck[x],bson.objectid.ObjectId)
                    
                elif x!=6:
                    assert isinstance(getCheck[x],str)
            elif getCheck[x] ==None:
                assert output[x] == False
        db.bookCollection.delete_one(condition)


class TestExecute:
    def test_executeUpdate(self,execute,filePathNSheet,db,finding):
        output = execute.executeUpdate()
        assert isinstance(output,list)
        assert len(output)==3
        assert isinstance(output[1],tuple)
        assert isinstance(output[2],tuple)
        pathNName,sheetName=filePathNSheet
        _,words = finding.getExcelWords(pathNName,sheetName)
        mySheet = finding.getSheet(pathNName,sheetName)
        idSheetColNRow = finding.findValueInExcel(finding.meaningIdTitleEx,pathNName,sheetName)
        columnId = idSheetColNRow[0][1]
        wordsMongo = finding.getWordsMongo()
        for eachWord in words:
            if eachWord in wordsMongo:
                if output[0] is True:
                    # test word
                
                    wordTypeEx,articleEx,voiceNameEx,voiceTelLinkEx,meaningIdEx= finding.getWordsFromExcelAllContent(eachWord)
                    # wordTypeEx,articleEx,voiceNameEx,voiceTelLinkEx,meaningIdEx = outputGetWord[0],outputGetWord[1],outputGetWord[2],outputGetWord[3],outputGetWord[4]
                    # wordType,article,voiceName,voiceTelLink,meaningId
                    outputWordContentMongo = finding.getWordsFromMongoAllContent(eachWord)
                    wordTypeMongo,articleMongo,voiceNameMongo,voiceTelLinkMongo,meaningIdMongo=outputWordContentMongo[0],outputWordContentMongo[1],outputWordContentMongo[2],outputWordContentMongo[3],outputWordContentMongo[4]
                    # print(f"{wordTypeMongo = }\n{articleMongo = }\n{voiceNameMongo = }\n{voiceTelLinkMongo = }\n{meaningIdMongo = }")
                    assert wordTypeEx == wordTypeMongo
                    assert articleEx == articleMongo
                    assert voiceNameEx == voiceNameMongo
                    assert voiceTelLinkEx == voiceTelLinkMongo
                    # print(f"form test word {meaningIdEx = }")
                    for eachOneIndex in range(len(meaningIdEx)):
                        assert meaningIdEx[eachOneIndex] == meaningIdMongo[eachOneIndex]
                if len(output[1]) !=0:
                # test meaning 
                    wordSheetColNRow = finding.findValueInExcel(eachWord,pathNName,sheetName)
                    assert len(wordSheetColNRow) == len(output[1])
                    for x in range(len(wordSheetColNRow)):
                        row = wordSheetColNRow[x][2]
                        id = mySheet.cell(row,columnId).value
                        meaningIdMongo,meaningNumDAFMongo,deutschMongo,englishMongo,synonymMongo,persianMongo,nameDeAudioMongo,addTelDeuMongo,nameSynAudioMongo,addTelSynMongo,nameEnAudioMongo,addTelEnMongo,namePerAudioMongo,addTelPerMongo = finding.getMeaningWordFromMongoAllDetails(eachWord,ObjectId(id))

                        meaningIdEx,meaningNumDAFEx,deutschEx,englishEx,synonymEx,persianEx,nameDeAudioEx,addTelDeuEx,nameSynAudioEx,addTelSynEx,nameEnAudioEx,addTelEnEx,namePerAudioEx,addTelPerEx= finding.getMeaningWordFromExcelAllDetails(eachWord,meaningId= id)
                        assert isinstance(meaningIdEx,list)
                        for x in range(len(meaningIdEx)):
                            assert  meaningIdMongo  == ObjectId(meaningIdEx[x])
                        assert  meaningNumDAFMongo == meaningNumDAFEx
                        assert  deutschMongo == deutschEx
                        assert  englishMongo == englishEx
                        assert  synonymMongo == synonymEx
                        assert  persianMongo == persianEx
                        # print(f"{nameDeAudioMongo = }\n{nameDeAudioEx = }")
                        assert  nameDeAudioMongo == nameDeAudioEx
                        assert  addTelDeuMongo == addTelDeuEx
                        assert  nameSynAudioMongo == nameSynAudioEx
                        assert  addTelSynMongo == addTelSynEx
                        assert  nameEnAudioMongo == nameEnAudioEx
                        assert  addTelEnMongo == addTelEnEx
                        assert  namePerAudioMongo == namePerAudioEx
                        assert  addTelPerMongo == addTelPerEx
                if len(output[2]) !=0:
                    #test book content
                    wordSheetColNRow = finding.findValueInExcel(eachWord,pathNName,sheetName)
                    assert len(wordSheetColNRow) == len(output[2])
                    for x in range(len(wordSheetColNRow)):
                        # bookName,chapter,content,page,publisher,meaningId 
                        bookNameEx,chapterEx,contentEx,pageEx,publisherEx,meaningIdEx = finding.getBookAllcontentFromExcel(eachWord)
                        # output ... bookName,chapter,content,page,publisher,meaningId
                        outputFunc =finding.getBookAllcontentFromMongo(eachWord)
                        if len(outputFunc) !=0:
                            
                            bookNameMongo,chapterMongo,contentMongo,pageMongo,publisherMongo,meaningIdMongo = outputFunc[0],outputFunc[1],outputFunc[2],outputFunc[3],outputFunc[4],outputFunc[5]
                            assert bookNameEx == bookNameMongo
                            assert chapterEx == chapterMongo
                            assert contentEx == contentMongo
                            assert pageEx == pageMongo
                            assert publisherEx == publisherMongo
                            assert meaningIdEx == meaningIdMongo
    @pytest.mark.parametrize("word,deutsch",[
        ("tschau","tschau! Ciao"),
        ("test2","deutsch2"),
        ("test3","deutsch3")

    ])
    def test_executeEnter(self,word,deutsch,execute,finding,db,filePathNSheet):
        # output = finding.executeEnter()
        # if output :
        pathNName,sheetName=filePathNSheet
        _,words = finding.getExcelWords(pathNName,sheetName)
        wordsMongo = finding.getWordsMongo()
        remainWords = set(words)-set(wordsMongo)
        # for eachWord in wordsMongo:

        outputExecuteEnter=execute.executeEnter(word,deutsch)
        assert isinstance(outputExecuteEnter,list)
        assert len(outputExecuteEnter) == 3
        # outputMeaningContents=entering.enterMeaningWordAllContents(word,deutsch)
        meaningVerfication,bookVerification,wordVerification = False,False,False
        meaningVerfication2,bookVerification2,wordVerification2 = False,False,False
        meaningIdFind,bookIdFind,wordIdFind = None,None,None
        if outputExecuteEnter[0] is not False:
            
        # for eachId in outputExecuteEnter:
            assert isinstance(outputExecuteEnter[0],bson.objectid.ObjectId)

            conditionMeaning = {"_id":outputExecuteEnter[0],"word":word,"deutsch":deutsch}
            for col in db.meaningCollection.find(conditionMeaning,{"_id":1}):
                for keys in col.keys():
                    meaningVerfication = True
                    meaningIdFind = col[keys]
            assert outputExecuteEnter[0]==meaningIdFind
        else:
            assert isinstance(outputExecuteEnter[0],bool)
            assert outputExecuteEnter[0] is False
            conditionMeaning = {"word":word,"deutsch":deutsch}
            for col in db.meaningCollection.find(conditionMeaning,{"_id":1}):
                for keys in col.keys():
                    # meaningVerfication = True
                    meaningIdFind = col[keys]
        
        # 
        # outputWordContents  = entering.enterMongoWordAllContents(word,outputExecuteEnter)
        if outputExecuteEnter[1] is not False:
            assert isinstance(outputExecuteEnter[1],bson.objectid.ObjectId)
            # print(f"{outputExecuteEnter[1] = }")
            conditionWord = {"_id":outputExecuteEnter[1],"word":word,"meaningId":{"$in":[outputExecuteEnter[0]]}}
            for col in db.wordCollection.find(conditionWord,{"_id":1}):
                for keys in col.keys():
                    wordVerification = True
                    wordIdFind = col[keys]
            assert outputExecuteEnter[1]==wordIdFind
        else:
            assert isinstance(outputExecuteEnter[1],bool)
            assert outputExecuteEnter[1] is False
        # outputBookContents = entering.enterMongoBookAllContents(word,outputExecuteEnter)
        if outputExecuteEnter[2] is not False:
            assert isinstance(outputExecuteEnter[2],bson.objectid.ObjectId)
            conditionBook = {"_id":outputExecuteEnter[2],"word":word,"meaningId":outputExecuteEnter[0]}
            for col in db.bookCollection.find(conditionBook,{"_id":1}):
                for keys in col.keys():
                    bookVerification = True
                    bookIdFind = col[keys]
            assert outputExecuteEnter[2]==bookIdFind
        else:
            assert isinstance(outputExecuteEnter[2],bool)
            assert outputExecuteEnter[2] is False
        if meaningVerfication is True:
            conditionMeaning = {"_id":outputExecuteEnter[0],"word":word,"deutsch":deutsch}
            db.meaningCollection.delete_one(conditionMeaning)
            for col in db.meaningCollection.find(conditionMeaning,{"_id":1}):
                for keys in col.keys():
                    meaningVerfication2=True
            assert meaningVerfication2 is False

        if wordVerification is True:
            conditionWord = {"_id":outputExecuteEnter[1],"word":word,"meaningId":{"$in":[outputExecuteEnter[0]]}}
            db.wordCollection.delete_one(conditionWord)
            for col in db.wordCollection.find(conditionWord,{"_id":1}):
                for keys in col.keys():
                    wordVerification2=True
            assert wordVerification2 is False

        if bookVerification is True:
            # print(f"{outputExecuteEnter = }")
            conditionBook = {"_id":outputExecuteEnter[2],"word":word,"meaningId":outputExecuteEnter[0]}
            db.bookCollection.delete_one(conditionBook)
            for col in db.bookCollection.find(conditionBook,{"_id":1}):
                for keys in col.keys():
                    bookVerification2=True
            assert bookVerification2 is False    

class TestSearch:
    @pytest.mark.parametrize("word,deutsch",[
        ("sich freuen [vr]","freu·en; freute, hat gefreut; [Vr] 1 sich (über etwas (Akk)) freuen wegen etwas ein Gefühl der Freude empfinden <sich sehr, ehrlich, riesig freuen>: sich über ein Geschenk, einen Anruf freuen; Ich habe mich sehr darüber gefreut, dass wir uns endlich kennen gelernt haben; Ich freue mich, Sie wieder zu sehen"),
        ("die Entschuldigung;-,en",None),
        (None,"Na·me der; -ns, -n; 1 das Wort (oder die Wörter), unter dem man eine Person oder Sache kennt und durch das man sie identifizieren kann <jemandem / etwas einen Namen geben; einen Namen für jemanden / etwas suchen, finden, aussuchen; jemandes Namen tragen; sich einen anderen Namen beilegen, zulegen; seinen Namen nennen, sagen, angeben, verschweigen>: Jeder nennt sie Nini, aber ihr wirklicher Name ist Martina; Sein Name ist Meier  || K-: Namen(s)änderung, Namen(s)verzeichnis, Namen(s)wechsel  || -K: Familienname, Firmenname, Flussname, Frauenname, Hundename, Jungenname, Künstlername, Ländername, Mädchenname, Männername, Ortsname, Städtename, Stoffname, Tiername, Vorname"),
        (None,None),
        ("someWord",None),
        (None,"someDeutsch"),
        ("someWord","someDeutsch")


    ])
    def test_simpleSearch(self,word,deutsch,db,search):
        # output = [{"wordId":wordId,"word":word,"wordType":wordType,"article":article,"wordVoiceName":wordVoiceName,"wordTelAdd":wordTelAdd,
        # meaning1:{
            # "meaningId":meaningId,"deutschMeaning":deutschMeaning,"deutschFileName":deutschFileName,"deutschTelAdd":deutschTelAdd,"englishMeaning":englishMeaning,"englishTelAdd":englishTelAdd,"synonymMeaning":synonymMeaning,"synonymFileName":synonymFileName,"synonymTelAdd":synonymTelAdd,"persianMeaning":persianMeaning,"persianFileName":persianFileName,"persianTelAdd":persianTelAdd,book1:{
            # "bookId":bookId,"bookName":bookName,"chapter":chapter,"content":content,"page":page,"lesson":lesson,"publisher":publisher
            # }

        output = search.simpleSearch(word,deutsch)
        if output is not False:
            strWords = ["wordId","word","wordType","article","wordVoiceName","wordtelegramLink","wordNumRow","meaning1","meaning2","meaning3","meaning4","meaning5","meaning6","meaning7","meaning8","meaning9","meaning10"]
            
            strMeaning=["meaningId","deutschMeaning","deuName","deuTelegramLink","englishMeaning","engName","engTelegramLink","synonymMeaning","synName","synTelegramLink","persianMeaning","meaningNumDaf","perName","perTelegramLink","meaningNumRow","book1","book2","book3","book4","book5","book6","book7","book8","book9","book10"]
            strBook=["bookId","bookName","chapter","content","page","lesson","publisher","bookNumRow"]
            objectIdFields = ["wordId","meaningId","bookId"]
            # bookNumRow,meaningNumRow,wordNumRow
            assert isinstance(output,list)
            assert len(output)>0
            wordNum = 0
            meaningNum = 0
            bookNum = 0
            for x in output:
                assert isinstance(x,dict) 
                assert x["wordNumRow"]
                if x["wordNumRow"]:
                    assert x["wordNumRow"] == wordNum+meaningNum+bookNum+1
                    assert isinstance(x["wordNumRow"],int)
                    wordNum=wordNum+1

                for field,val in x.items():

                    assert field in strWords
                    if field in objectIdFields:
                        assert isinstance(val,bson.objectid.ObjectId)
                    for i in range(1,10):
                        if field == f"meaning{i}":
                            assert val["meaningNumRow"]
                            if val["meaningNumRow"]:
                                assert val["meaningNumRow"] == wordNum+meaningNum+bookNum+1
                                assert isinstance(val["meaningNumRow"],int)
                                meaningNum=meaningNum+1
                            for fieldMeaning,valmeaning in val.items():
                                # print(f"fieldMeaning = {fieldMeaning}")
                                assert fieldMeaning in strMeaning 
                                if fieldMeaning in objectIdFields:
                                    assert isinstance(valmeaning,bson.objectid.ObjectId) 
                                for j in range(1,10):
                                    if fieldMeaning == f"book{j}":
                                        assert valmeaning["bookNumRow"]
                                        if valmeaning["bookNumRow"]:
                                            assert valmeaning["bookNumRow"] == wordNum+meaningNum+bookNum+1
                                            assert isinstance(valmeaning["bookNumRow"],int)
                                            meaningNum=meaningNum+1
                                        for fieldBook,valBook in valmeaning.items():
                                            if fieldBook == "bookNumRow":
                                                isinstance(valBook,int)
                                                print(f"wordNum = {wordNum}\nmeaningNum = {meaningNum}\nbookNum = {bookNum}")
                                                assert valBook == wordNum+meaningNum+bookNum
                                                bookNum = bookNum+1 
                                            # print(f"valmeaning = {valmeaning}")
                                            # print(f"wordId = {x['wordId']}\nmeaningId = {val['meaningId']}\nbookId = {valmeaning['bookId']}")
                                            assert valmeaning["bookNumRow"]
                                            assert fieldBook in strBook 
                                            if fieldBook in objectIdFields:
                                                assert isinstance(valBook,bson.objectid.ObjectId) 


class TestAdding:
    def test_addWordDetailsToMeaningFromMongoToMongo(self,adding,db):
        output = adding.addWordDetailsToMeaningFromMongoToMongo()
        assert isinstance(output,list)
        for x in output:
            assert len(x)==4
            assert isinstance(x[0],bson.objectid.ObjectId)
            assert isinstance(x[1],str)
            assert isinstance(x[2],str)
            assert isinstance(x[3],str)
            verification = False
            # condition = {"wordDetails":{"$elemMatch":{"wordId":x[0],"word":x[1],"wordType":x[2],"article":x[3]}}
            condition = {"wordDetails.wordId":x[0]}
            
            for col in db.meaningCollection.find(condition,{"_id":1}):
    
                for keys in col.keys():
                    verification = True
            assert verification is True



    def test_addMeaningIdToWordCollecFromMongoToMongo(self,db,adding):
        #without having even ids
        output = adding.addMeaningIdToWordCollecFromMongoToMongo()
        assert isinstance(output,list)
        for i in output:
            assert isinstance(i,tuple)
            assert len(i) == 2
            assert isinstance(i[0],str)
            assert isinstance(i[1],list)
            for j in i[1]:
                assert isinstance(j,bson.objectid.ObjectId)
                assert j !=""

        meaningWords = []
        meaningIdsInmeaningCollec = []
        
        for col in db.meaningCollection.find({},{"_id":1,"word":1}):
            for field,val in col.items():
                if field == "word":
                    meaningWords.append(val)
                elif field == "_id":
                    meaningIdsInmeaningCollec.append(val)
        words = []
        meaningidsInWordCol = []
        wordMeaningIdNword = []
        for col in db.wordCollection.find({},{"_id":0,"word":1,"meaningId":1}):
            temp = []
            for field,val in col.items():
                if field == "word":
                    temp.append(val)
                    words.append(val)
                elif field == "meaningId":

                    temp.append(val)
                    meaningidsInWordCol=meaningidsInWordCol + val
            wordMeaningIdNword.append(tuple(temp))
        meaningidsInWordColWithoutEmpty = [x for x in meaningidsInWordCol if x!=""] 
        for i in    meaningidsInWordColWithoutEmpty:
            assert i !=""  
        assert len(list(set(meaningidsInWordCol))) ==len(meaningidsInWordCol)
        assert len(words) == len(meaningWords)
        assert len(meaningidsInWordCol) == len(meaningidsInWordColWithoutEmpty)
        # assert len(words) 
        for i in wordMeaningIdNword:
            for j in i[1]:
                veification = False
                for col in db.meaningCollection.find({"_id":j},{"word":1,"_id":0}):
                    for field,val in col.items():
                        assert i[0]==val
                        veification =True
                assert veification is True

      
class TestEditing:
    def test_DeleteMeaningAndWordIdFromBook(self,db):
        db.bookCollection.update_many( {},{ "$unset": {"meaningAndWordId": ""}})
    @pytest.mark.parametrize("newField,beforeField",[
        ("meaningId","meaningid"),
        ("meaningId","meaningId")

    ])
    def test_editFieldInCollection(self,editing,db,newField,beforeField):   
        
        output = editing.editingWordColField(newField,beforeField)
        assert isinstance(output,bool)
        if output is True:
            verficationNewField = False
            verificationBeforeField = False
            condition ={} 
            for col in db.wordCollection.find(condition,{"_id":1,newField:1,beforeField:1}):
                for field,val in col.items():
                    if field == beforeField:
                        verificationBeforeField = True
                    if field == newField:
                        verficationNewField = True
            assert verificationBeforeField is False
            assert verficationNewField is True

    def test_meaningIdAndWordIdTomeaningId(self,db):
        bookName,chapter,content,page,lesson,publisher,meaningId,word=None,None,None,None,None,None,None,None
        outputMongo = {"bookName":1,"chapter":1,"content":1,"page":1,"lesson":1,"publisher":1,"meaningId":1,"word":1,"_id":0}
        for col in db.bookCollection.find({},outputMongo):
            for field,val in col.items():
                if field == "bookName":
                    bookName = val
                if field == "chapter":
                    chapter = val
                if field == "content":
                    content = val
                if field == "page":
                    page = val
                if field == "lesson":
                    lesson = val
                if field == "publisher":
                    publisher = val
                # if field =="meaningId":
                #     meaningId = val
                if field == "word":
                    word = val

            input = {"bookName":bookName,"chapter":chapter,"content":content,"page":page,"lesson":lesson,"publisher":publisher,"meaningId":meaningId,"word":word}
            # db.wordCollection.update_many(condition,{"$set":input})
            condition = {"word":word}
            db.bookCollection.update_many(condition,{"$set":input})
        # db.bookCollection.update_many( {},{ "$unset": {"meaningAndWordId": ""}})

    def test_addmeaningIdToBookCollection(self,db):
        outputMongo = {"_id":1,"word":1}
        wordNId = []
        for col in db.meaningCollection.find({},outputMongo):
            temp = []
            for field,val in col.items():
                if field == "_id":
                    temp.append(val)
                if field == "word":
                    temp.append(val)
            wordNId.append(tuple(temp))

        for i in wordNId:
            print(f"{wordNId = }")
            db.bookCollection.update_many({"word":i[1]},{"$set":{"meaningId":i[0]}})

    def test_deleteBookIdWordIdFrommeaningId(self,db):
        db.meaningCollection.update_many( {},{ "$unset": {"wordId": ""}})
        db.meaningCollection.update_many( {},{ "$unset": {"bookId": ""}})

    def test_deleteAllVoice(self,db):
        db.meaningCollection.update_many( {},{ "$unset": {"allVoice": ""}})

    @pytest.mark.parametrize("wordId,word,wordType,article,voiceName,telegramLink",[
        ("5f30077014082a0f00465048","testWord1","wordType1","article1","voiceName1","telegramLink1"),
        ("5f30077014082a0f00465048","testWord2","wordType2","article2","voiceName2","telegramLink2"),
        ("5f30077014082a0f00465048","testWord3","wordType3","article3","voiceName3","telegramLink3"),
        ("5f30077014082a0f00465048","testWord4","wordType4","article4","voiceName4","telegramLink4"),
        ("5f30077014082a0f00465048","testWord5","wordType5","article5","voiceName5","telegramLink5")
    ])
    def test_editWord(self,db,editing,wordId,word,wordType,article,voiceName,telegramLink):
        wordId,word,wordType,article,voiceName,telegramLink,meaningId=enterAWord 
        output= editing.editWord(wordId,word,wordType,article,voiceName,telegramLink)
        if output is True:
            condition = {"_id":ObjectId(wordId)}
            outputMongo = {"word":1,"article":1,"wordType":1,"voice":1}
            wordMongo,wordTypeMongo,articleMongo,voiceNameMongo,telegramLinkMongo = None,None,None,None,None
            for col in db.wordCollection(condition,outputMongo):
                for field,val in col.items():
                    if field =="word":
                        wordMongo = val
                    elif field == "article":
                        articleMongo = val
                    elif field == "wordType":
                        wordTypeMongo = val
                    elif field == "voice":
                        voiceNameMongo = val[name]
                        telegramLinkMongo = val[telegramLink]
            if word !=None:
                assert word == wordMongo
            if wordType !=None:
                assert wordType == wordTypeMongo
            if article != None:
                assert article == articleMongo
            if voiceName != None:
                assert voiceName == voiceNameMongo
            if telegramLink != None:
                assert telegramLink == telegramLinkMongo

class TestDeleting:
    def test_deleteWord(self,db,deleting,enterAWord):
        # print(enterAWord)
        wordId,word,wordType,article,voiceName,telegramLink,meaningId=enterAWord   
        output = deleting.deleteWord(wordId)
        verification = False
        if output is True:
            for col in db.wordCollection.find({"_id":ObjectId(wordId)},{"word":1}):
                for keys in col.keys():
                    verification = True
            assert verification is False







            

                
